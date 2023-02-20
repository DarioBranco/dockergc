import openpyxl
import csv
import sys
import glob


def calculateSetKPI(pilot,demon,year):
    years = year.split("-")
    files = glob.glob("/home/branco/public_html/gccalculator/csv/newCalculator/Utils/shared/other_data/*.xlsx")
    selectedFile = '0'
    print(files[0].split('/')[-1].split("-")[0])
    for file in files:
        if(file.split('/')[-1].split("-")[0] == pilot and file.split('/')[-1].split("-")[1] == demon):
            selectedFile = file
    print(selectedFile)
    wb_obj = openpyxl.load_workbook(selectedFile)

    sheet = wb_obj['G5.1']
    KPI511 = [[]]
    KPI512 = [[]]
    KPI513 = [[]]
    KPI514 = [[]]
    KPI515 = [[]]
    for row in sheet.iter_rows(max_row = sheet.max_row):
               if(str(row[2].value) in years):
                    KPI511[0].append(float(row[4].value))
                    KPI512[0].append(float(row[5].value))
                    KPI513[0].append(float(row[6].value))
                    KPI514[0].append(float(row[7].value))
                    KPI515[0].append(float(row[8].value))

    KPI521 = [[]]
    KPI522 = [[]]
    KPI523 = [[]]
    KPI524 = [[]]
    KPI525 = [[]]
    KPI526 = [[]]
    sheet = wb_obj['G5.2']

    for row in sheet.iter_rows(max_row = sheet.max_row):
               if(str(row[2].value) in years):
                    KPI521[0].append(float(row[4].value))
                    KPI522[0].append(float(row[5].value))
                    KPI523[0].append(float(row[6].value))
                    KPI524[0].append(float(row[7].value))
                    KPI525[0].append(float(row[8].value))
                    KPI526[0].append(float(row[9].value))


    sheet = wb_obj['G5.7']
    KPI57 = [[]]
    for row in sheet.iter_rows(max_row = sheet.max_row):
               if(str(row[2].value) in years):
                     try:
                        KPI57[0].append(float(row[4].value))
                     except:
                        KPI57[0].append([0])


    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data511.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI511)	
	
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data512.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI512)	
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data513.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI513)	
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data514.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI514)
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data515.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI515)   
	   
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data521.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI521)	
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data522.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI522)	
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data523.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI523)
    file = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data524.csv", "w")
    with file:
       write = csv.writer(file)
       write.writerows(KPI524)   	   
	         
    file2 = open("/home/branco/public_html/gccalculator/csv/kpi/business/"+pilot+"_"+demon+"_"+year+"_data1.csv","w")
    with file2:
       write = csv.writer(file2)
       write.writerows(KPI57)



if __name__ == "__main__" :

      """pilot=sys.argv[1]
      demon=sys.argv[2]
      years = sys.argv[3].split("-")
      user = sys.argv[4]
      calculateSetKPI("OSLO","D1","2019")
      calculateSetKPI("OSLO","D1","2020")
      calculateSetKPI("OSLO","D1","2021")
      calculateSetKPI("OSLO","D1","2019-2020")
      calculateSetKPI("OSLO","D1","2020-2021")
      calculateSetKPI("OSLO","D1","2019-2020-2021")
      calculateSetKPI("OSLO","D2","2019")
      calculateSetKPI("OSLO","D2","2020")
      calculateSetKPI("OSLO","D2","2021")
      calculateSetKPI("OSLO","D2","2019-2020")
      calculateSetKPI("OSLO","D2","2020-2021")
      calculateSetKPI("OSLO","D2","2019-2020-2021")
      calculateSetKPI("OSLO","D3","2019")
      calculateSetKPI("OSLO","D3","2020")
      calculateSetKPI("OSLO","D3","2021")
      calculateSetKPI("OSLO","D3","2019-2020")
      calculateSetKPI("OSLO","D3","2020-2021")
      calculateSetKPI("OSLO","D3","2019-2020-2021")"""
      calculateSetKPI("Barcelona","D1","2019")
      calculateSetKPI("Barcelona","D1","2020")
      calculateSetKPI("Barcelona","D1","2021")
      calculateSetKPI("Barcelona","D1","2019-2020")
      calculateSetKPI("Barcelona","D1","2020-2021")
      calculateSetKPI("Barcelona","D1","2019-2020-2021")
      """
      calculateSetKPI("BARCELONA","D2","2019")
      calculateSetKPI("BARCELONA","D2","2020")
      calculateSetKPI("BARCELONA","D2","2021")
      calculateSetKPI("BARCELONA","D2","2019-2020")
      calculateSetKPI("BARCELONA","D2","2020-2021")
      calculateSetKPI("BARCELONA","D2","2019-2020-2021")
      calculateSetKPI("BARCELONA","D3","2019")
      calculateSetKPI("BARCELONA","D3","2020")
      calculateSetKPI("BARCELONA","D3","2021")
      calculateSetKPI("BARCELONA","D3","2019-2020")
      calculateSetKPI("BARCELONA","D3","2020-2021")
      calculateSetKPI("BARCELONA","D3","2019-2020-2021")
      calculateSetKPI("BREMEN","D1","2019")
      calculateSetKPI("BREMEN","D1","2020")
      calculateSetKPI("BREMEN","D1","2021")
      calculateSetKPI("BREMEN","D1","2019-2020")
      calculateSetKPI("BREMEN","D1","2020-2021")
      calculateSetKPI("BREMEN","D1","2019-2020-2021")
      calculateSetKPI("BREMEN","D2","2019")
      calculateSetKPI("BREMEN","D2","2020")
      calculateSetKPI("BREMEN","D2","2021")
      calculateSetKPI("BREMEN","D2","2019-2020")
      calculateSetKPI("BREMEN","D2","2020-2021")
      calculateSetKPI("BREMEN","D2","2019-2020-2021")
      calculateSetKPI("BREMEN","D3","2019")
      calculateSetKPI("BREMEN","D3","2020")
      calculateSetKPI("BREMEN","D3","2021")
      calculateSetKPI("BREMEN","D3","2019-2020")
      calculateSetKPI("BREMEN","D3","2020-2021")
      calculateSetKPI("BREMEN","D3","2019-2020-2021")
	   """

	
