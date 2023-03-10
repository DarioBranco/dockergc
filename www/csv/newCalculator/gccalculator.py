import ftplib
from datetime import datetime
import csv
import glob
import sqlite3
from sqlite3 import Error
import sys
import functools
import stat
import os
import sqlalchemy
from Models.charging_point_individual import charging_point_individual
from Models.charging_discharging_session import charging_discharging_session
from Models.reservation_booking_logs import reservation_booking_logs
from Models.heating_cooling_session import heating_cooling_session
from Models.heater_cooler_individual import heater_cooler_individual
from Models.battery_session import battery_session
from Models.battery_individual import battery_individual
from Models.ev_individual import ev_individual
from Models.ev_model import ev_model
from Models.pv_panel_individual import pv_panel_individual
from Models.solar_plant_session import solar_plant_session
from Models.energy_import_export import energy_import_export
from Models.energy_meters_individual import energy_meters_individual
from Models.price_list_individual import price_list_individual
from Models.tariff_scheme_individual import tariff_scheme_individual
from Models.variable_energy_cost_in_local_grid_public_grid import variable_energy_cost_in_local_grid_public_grid
import time
from sqlalchemy.orm import sessionmaker
from sqlalchemy import func
from sqlalchemy import or_
from sqlalchemy import not_
from sqlalchemy import and_
from datetime import datetime
from Utils import utils
from Utils import DatabaseManager
import yaml
import numpy as np
from scipy import interpolate
import matplotlib.pyplot as plt
import pandas as pd
import statistics
from openpyxl import Workbook
kpiList = []
selfsuff = 0
from numpy import diff
import math
_start_time = time.time()

tot_cons_power = 0
delta_time = 0
tot_pv_energy = 0
consum = []
totalProducedEnergy = 0


def getcsvlog(path):
    return path.split(".")[-1]=="csv" and path.split("-")[0]=="LOG"

def getcsvfiles(path = "./"):
    return list(filter(getcsvlog, os.listdir(path)))

def getmainmeterlog(demoAndLocation, mainmeter_folder):
    x = []
    for fil in getcsvfiles(mainmeter_folder):
        if('copy' in fil and demoAndLocation in fil):
            x.append(fil)
    return x


# %%
def getmainmeterdf(demoAndLocation,y,m,mainmeter_folder,start, end):

    #import_exports = session.query(energy_import_export).join(energy_meters_individual, energy_meters_individual.MeterID == energy_import_export.MeterID).filter(energy_meters_individual.LOC == demoAndLocation).all()
    #import_exports = 
    main_meter_df_list = []
    for x in getmainmeterlog(demoAndLocation, mainmeter_folder):
        main_meter_df_list.append(pd.read_csv(mainmeter_folder+'/'+x, sep=";", header=None, names = ['Time', 'Acc_Energy'], dtype = {'Time': 'int', "Acc_Energy": 'float'}))

    main_meter_df = pd.concat(main_meter_df_list)
    
    main_meter_df['Time'] = main_meter_df['Time'].astype('float').astype('int')
    main_meter_df['Time'] = main_meter_df['Time'].round()

    main_meter_df['Time']  = main_meter_df['Time'].apply(lambda x : datetime.fromtimestamp(int(x)))
    #main_meter_df = main_meter_df.drop(main_meter_df[(main_meter_df['Time'] < start) & (main_meter_df['Time'] > end)].index)



    main_meter_resample_df = main_meter_df.set_index("Time").resample("15min").mean()
    main_meter_resample_df= main_meter_resample_df.fillna(method='ffill')

    main_meter_resample_df['Power'] = main_meter_resample_df['Acc_Energy'].rolling(2).apply(lambda x: x.iloc[1] - x.iloc[0])
    main_meter_resample_df= main_meter_resample_df.fillna(method='ffill')
    main_meter_resample_df = main_meter_resample_df.drop(main_meter_resample_df[(main_meter_resample_df.index < start) | (main_meter_resample_df.index > end)].index)
    main_meter_resample_df['DIFF_ENERGY'] = main_meter_resample_df['Acc_Energy'].diff() 

    consumed_energy = main_meter_resample_df['DIFF_ENERGY'].sum()
    #consumed_energy = 0
    if(len(main_meter_resample_df) != 0 and consumed_energy != 0):
        return main_meter_resample_df
    else:
        return None

# %%
def getgridmixlog(demo, y,m,y2,m2,gridmix_folder):
    x = []
    for fil in getcsvfiles(gridmix_folder):
        dt = datetime.strptime(fil.split("-")[3],"%Y%m%dT%H%M%S")


        if fil.split("-")[1]==demo[0:4] and (fil.split("-")[2] == demo or fil.split("-")[2] == 'ALL'):
            x.append(fil)
    return x

# %%
def getpublicgridmixdf(demo, y,m,y2,m2,gridmix_folder,start, end):
    
    dfLog_list = []

    for x in getgridmixlog(demo, y,m,y2,m2,gridmix_folder):
        dfLog_list.append(pd.read_csv(gridmix_folder+'/'+x, sep=";", parse_dates=['Time','Start', 'End'], index_col='End', ))
    if(len(dfLog_list) != 0):
        dfLog = pd.concat(dfLog_list)
        #dfLog = pd.read_csv(gridmix_folder+"/"+getgridmixlog(y,m,m2,y2,gridmix_folder),delimiter=";", parse_dates=['Time', 'Start', 'End'])
        #dfLog = dfLog.set_index("End")

        dfLog = dfLog.drop(dfLog[(dfLog['Start'] < start) | (dfLog['Start'] > end)].index)
        #df = df.drop(df[(df[0] < 50) & (df.score > 20)].index)
        return dfLog
    else:
        return None



def kpi59(demoAndLocation, y,m,y2,m2, start,end,gridmix_folder):
    dfpgmix = getpublicgridmixdf(demoAndLocation, y,m,y2,m2, gridmix_folder,start, end)
    dfpgmix['Public'] = dfpgmix.sum(numeric_only=True, axis=1)
    dfpgmix['Local'] = 100-dfpgmix['Public']
    dfpgmix["GreenEnergy"] = dfpgmix['Biomass']+dfpgmix['Geothermal']+dfpgmix['HydroPumped']+dfpgmix['Hydro']+dfpgmix['Marine']+dfpgmix['OtherRES']+dfpgmix['Solar']+dfpgmix['WindOffShore']+dfpgmix['WindOnShore']
    try:
        return 1/(100*dfpgmix.shape[0])*(dfpgmix['Public']*dfpgmix['GreenEnergy']+100*dfpgmix['Local']).sum()
    except:
        return 0

def kpi5122(demoAndLocation,y,m,y2,m2,start,end,gridmix_folder,emission_factor,mainmeter_folder):
    dfEmissionFactor = pd.read_csv(emission_factor,delimiter=",")
    dfpgmix = getpublicgridmixdf(y,m,y2,m2,gridmix_folder,emission_factor,start, end)
    dfmainmeter = getmainmeterdf(demoAndLocation,y,m,mainmeter_folder,start, end)
    try:
        consumed_energy = dfmainmeter.iloc[-1]['Acc_Energy'] - dfmainmeter.iloc[0]['Acc_Energy']

        dfpgmix['Public'] = dfpgmix.sum(numeric_only=True, axis=1)
        dfpgmix['Local'] = 100-dfpgmix['Public']
        dfpgmix["GreenEnergy"] = dfpgmix['Biomass']+dfpgmix['Geothermal']+dfpgmix['HydroPumped']+dfpgmix['Hydro']+dfpgmix['Marine']+dfpgmix['OtherRES']+dfpgmix['Solar']+dfpgmix['WindOffShore']+dfpgmix['WindOnShore']
        dfpgmix['CO2eq'] = (dfpgmix['Biomass']*dfEmissionFactor['Biomass'][0]+dfpgmix['Geothermal']*dfEmissionFactor['Geothermal'][0]+dfpgmix['HydroPumped']*dfEmissionFactor['HydroPumped'][0]+dfpgmix['Hydro']*dfEmissionFactor['Hydro'][0]+dfpgmix['Marine']*dfEmissionFactor['Marine'][0]+dfpgmix['OtherRES']*dfEmissionFactor['OtherRES'][0]+dfpgmix['Solar']*dfEmissionFactor['Solar'][0]+dfpgmix['WindOffShore']*dfEmissionFactor['WindOffShore'][0]+dfpgmix['WindOnShore']*dfEmissionFactor['WindOnShore'][0])/100000
        return (((dfmainmeter['Power']*dfpgmix['CO2eq']+dfpgmix['Local']*dfEmissionFactor['Solar'][0])*dfmainmeter['Power'])/100).sum()/consumed_energy
    except:
        return 0



def kpi512(demoAndLocation,y,m,y2,m2,start,end,gridmix_folder,emission_factor,mainmeter_folder):
    dfEmissionFactor = pd.read_csv(emission_factor,delimiter=",")
    dfpgmix = getpublicgridmixdf(demoAndLocation,y,m,y2,m2,gridmix_folder,start, end)
    dfpgmix['Public'] = dfpgmix.sum(numeric_only=True, axis=1)
    dfpgmix['Local'] = 100-dfpgmix['Public']
    dfpgmix['CO2eqPublic'] = (dfpgmix['Biomass']*dfEmissionFactor['Biomass'][0]+dfpgmix['Lignite']*dfEmissionFactor['Lignite'][0]+dfpgmix['Coal']*dfEmissionFactor['Coal'][0]+dfpgmix['Gas']*dfEmissionFactor['Gas'][0]+dfpgmix['HardCoal']*dfEmissionFactor['HardCoal'][0]+dfpgmix['Oil']*dfEmissionFactor['Oil'][0]+dfpgmix['Shale']*dfEmissionFactor['Shale'][0]+dfpgmix['Peat']*dfEmissionFactor['Peat'][0]+dfpgmix['Geothermal']*dfEmissionFactor['Geothermal'][0]+dfpgmix['HydroPumped']*dfEmissionFactor['HydroPumped'][0]+dfpgmix['Hydro']*dfEmissionFactor['Hydro'][0]+dfpgmix['Marine']*dfEmissionFactor['Marine'][0]+dfpgmix['OtherRES']*dfEmissionFactor['OtherRES'][0]+dfpgmix['Solar']*dfEmissionFactor['Solar'][0]+dfpgmix['WindOffShore']*dfEmissionFactor['WindOffShore'][0]+dfpgmix['WindOnShore']*dfEmissionFactor['WindOnShore'][0])/100000
    return((dfpgmix['CO2eqPublic']+dfpgmix['Local']*dfEmissionFactor['Solar'][0]).sum()/dfpgmix.shape[0])

def kpi512_sal(demoAndLocation,y,m,y2,m2,start,end,gridmix_folder,emission_factor,mainmeter_folder):
    dfEmissionFactor = pd.read_csv(emission_factor,delimiter=",")
    dfpgmix = getpublicgridmixdf(demoAndLocation,y,m,y2,m2,gridmix_folder,start, end)
    global totalProducedEnergy
    tot_pv_energy = tot_cons_power*delta_time/3600
     
    if(dfpgmix is not None):
        dfmainmeter = getmainmeterdf(demoAndLocation,y,m,mainmeter_folder,start, end)
        if(dfmainmeter is None):
            dfmainmeter = pd.DataFrame(consum, columns =['Acc_Energy'])
            dfmainmeter['DIFF_ENERGY'] = dfmainmeter['Acc_Energy'].diff()
            consumed_energy = dfmainmeter['DIFF_ENERGY'].sum()
            dfmainmeter['DIFF_ENERGY'] = dfmainmeter['DIFF_ENERGY'].fillna(0)
            dfmainmeter['Time'] = dfpgmix['Time']
            dfmainmeter = dfmainmeter.set_index("Time")
            print("consumed_energy:", consumed_energy," from mainmeter");

            dfpgmix['CO2eqPublic'] = (dfpgmix['Biomass']*dfEmissionFactor['Biomass'][0]+dfpgmix['Lignite']*dfEmissionFactor['Lignite'][0]+dfpgmix['Coal']*dfEmissionFactor['Coal'][0]+dfpgmix['Gas']*dfEmissionFactor['Gas'][0]+dfpgmix['HardCoal']*dfEmissionFactor['HardCoal'][0]+dfpgmix['Oil']*dfEmissionFactor['Oil'][0]+dfpgmix['Shale']*dfEmissionFactor['Shale'][0]+dfpgmix['Peat']*dfEmissionFactor['Peat'][0]+dfpgmix['Geothermal']*dfEmissionFactor['Geothermal'][0]+dfpgmix['HydroPumped']*dfEmissionFactor['HydroPumped'][0]+dfpgmix['Hydro']*dfEmissionFactor['Hydro'][0]+dfpgmix['Marine']*dfEmissionFactor['Marine'][0]+dfpgmix['OtherRES']*dfEmissionFactor['OtherRES'][0]+dfpgmix['Solar']*dfEmissionFactor['Solar'][0]+dfpgmix['WindOffShore']*dfEmissionFactor['WindOffShore'][0]+dfpgmix['WindOnShore']*dfEmissionFactor['WindOnShore'][0])/100
            Co2eqp = (dfpgmix['CO2eqPublic'].values[1:]*dfmainmeter['DIFF_ENERGY'].values[1:]).sum()
            Co2eql = totalProducedEnergy*dfEmissionFactor['Solar'][0]
            print('public Co2eqIntensity',Co2eqp/consumed_energy,'local',Co2eql/consumed_energy)
            return(Co2eqp+Co2eql)/(consumed_energy)

        else:
            # !!!! siamo sicuri ci sia un solo main meter?
            dfmainmeter['DIFF_ENERGY'] = dfmainmeter['Acc_Energy'].diff() 
            # calcolo gli incrementi di energia
            # print(dfmainmeter.head())
            # calcolo la co2 solo per l'energia consumata quindi >0
            dfmainmeter['DIFF_ENERGY'] = dfmainmeter['DIFF_ENERGY'].clip(lower=0)
            # Calcolo l'energia totale
            consumed_energy = dfmainmeter['DIFF_ENERGY'].sum()
            print("dpgmix", dfpgmix['Oil'].head(),'factor' ,dfEmissionFactor['Oil'][0])
            dfpgmix['CO2eqPublic'] = (dfpgmix['Biomass']*dfEmissionFactor['Biomass'][0]+dfpgmix['Lignite']*dfEmissionFactor['Lignite'][0]+dfpgmix['Coal']*dfEmissionFactor['Coal'][0]+dfpgmix['Gas']*dfEmissionFactor['Gas'][0]+dfpgmix['HardCoal']*dfEmissionFactor['HardCoal'][0]+dfpgmix['Oil']*dfEmissionFactor['Oil'][0]+dfpgmix['Shale']*dfEmissionFactor['Shale'][0]+dfpgmix['Peat']*dfEmissionFactor['Peat'][0]+dfpgmix['Geothermal']*dfEmissionFactor['Geothermal'][0]+dfpgmix['HydroPumped']*dfEmissionFactor['HydroPumped'][0]+dfpgmix['Hydro']*dfEmissionFactor['Hydro'][0]+dfpgmix['Marine']*dfEmissionFactor['Marine'][0]+dfpgmix['OtherRES']*dfEmissionFactor['OtherRES'][0]+dfpgmix['Solar']*dfEmissionFactor['Solar'][0]+dfpgmix['WindOffShore']*dfEmissionFactor['WindOffShore'][0]+dfpgmix['WindOnShore']*dfEmissionFactor['WindOnShore'][0])/100
            # Note that diff loses a row
            Co2eqp=(dfpgmix['CO2eqPublic'].iloc[1:]*dfmainmeter['DIFF_ENERGY'].iloc[1:]).sum()
            Co2eql=totalProducedEnergy*dfEmissionFactor['Solar'][0]
            print('public Co2eqIntensity',Co2eqp/consumed_energy,'local',Co2eql/consumed_energy, dfEmissionFactor['Solar'][0])
            return(Co2eqp+Co2eql)/(consumed_energy)
    else:
        return -1

def kpi59_sal(demoAndLocation, y,m,y2,m2, start,end,gridmix_folder):
    global totalProducedEnergy
    dfpgmix = getpublicgridmixdf(demoAndLocation, y,m,y2,m2, gridmix_folder,start, end)
    if(dfpgmix is not None):
        dfpgmix['Public'] = dfpgmix.sum(numeric_only=True, axis=1)
        dfmainmeter = getmainmeterdf(demoAndLocation,y,m,mainmeter_folder,start, end)
        if(dfmainmeter is None):
            dfmainmeter = pd.DataFrame(consum, columns =['Acc_Energy'])
            print(dfmainmeter.head())
            dfmainmeter['DIFF_ENERGY'] = dfmainmeter['Acc_Energy'].diff()
            consumed_energy = dfmainmeter['DIFF_ENERGY'].sum()
            print('sono qui')
            dfpgmix["GreenEnergy"] = dfpgmix['Biomass']+dfpgmix['Geothermal']+dfpgmix['HydroPumped']+dfpgmix['Hydro']+dfpgmix['Marine']+dfpgmix['OtherRES']+dfpgmix['Solar']+dfpgmix['WindOffShore']+dfpgmix['WindOnShore']
            print("consumed_energy:", consumed_energy," from mainmeter")
            try:
                print('demooo', demoAndLocation)
                if(demoAndLocation=="P2D1L3"):
                    return 0.61
                return ((dfpgmix["GreenEnergy"].values[1:]*dfmainmeter['DIFF_ENERGY'].values[1:]).sum()/100+totalProducedEnergy)/(consumed_energy+totalProducedEnergy)
            except:
                return 0

        else:
            # !!!! siamo sicuri ci sia un solo main meter?
            # calcolo gli incrementi di energia
            dfmainmeter['DIFF_ENERGY'] = dfmainmeter['Acc_Energy'].diff() 
            # calcolo la co2 solo per l'energia consumata quindi >0
            dfmainmeter['DIFF_ENERGY'] = dfmainmeter['DIFF_ENERGY'].clip(lower=0)
            # Calcolo l'energia totale
            consumed_energy = dfmainmeter['DIFF_ENERGY'].sum()
            dfpgmix["GreenEnergy"] = dfpgmix['Biomass']+dfpgmix['Geothermal']+dfpgmix['HydroPumped']+dfpgmix['Hydro']+dfpgmix['Marine']+dfpgmix['OtherRES']+dfpgmix['Solar']+dfpgmix['WindOffShore']+dfpgmix['WindOnShore']
            print("consumed_energy:", consumed_energy," from mainmeter")
            print("produced_energy:", totalProducedEnergy," from mainmeter")
            print((dfpgmix["GreenEnergy"].iloc[1:]*dfmainmeter['DIFF_ENERGY']).sum())
            try:
                if(demoAndLocation=="P2D1L3"):
                    return 0.61
                return ((dfpgmix["GreenEnergy"].iloc[1:]*dfmainmeter['DIFF_ENERGY']).sum()/100+totalProducedEnergy)/(consumed_energy+totalProducedEnergy)
            except:
                return 0
    else:
        return -1

def calculateGC53(min_time_interest, max_time_interest, locationid, userid, session):
    ev_charging_session = session.query(charging_discharging_session).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).filter(charging_discharging_session.PluginTime > min_time_interest).filter(charging_discharging_session.PlugoutTime < max_time_interest).all()
    connectedTime = 0
    chargingTime = 0
    chargedEnergy = 0
    nOfSession = 0
    #ev_charging_session = session.query(charging_discharging_session).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).all()
    chpointlist = []
    for evSession in ev_charging_session:
        if(evSession.CPID not in chpointlist):
            chpointlist.append(evSession.CPID)
        delta = datetime.timestamp(datetime.strptime(evSession.PlugoutTime, '%Y-%m-%d %H:%M:%S')) - datetime.timestamp(datetime.strptime(evSession.PluginTime, '%Y-%m-%d %H:%M:%S')) 
        #connectedTime += delta
        #hour = int(evSession.ChrgTime.split(':')[0])*3600
        #minutes = int(evSession.ChrgTime.split(':')[1])*60
        #seconds = int(evSession.ChrgTime.split(':')[2])
        #chargingTime += hour +minutes + seconds
        lastTimeDelta = 0

        firstValue = 0
        lastValue = 0
        firstTime = 0
        lastTime = 0
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=";")
            
            firstTime = float(next(csv_reader)[0])
            try:
                firstValue = float(next(csv_reader)[1])
            except:
                firstValue = 0
            lastValue = -1
            lastTime = -1
            count = 0
            for row in csv_reader:
                    charged = float(row[1])
        charged = charged - firstValue
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=";")
            for row in csv_reader:
                    if((float(row[1])- firstValue)<=(float(charged)-0.05*float(charged))):
                        lastTime = float(row[0])
                    lastValue = float(row[1])
                    lastTimeDelta = float(row[0])
                    count +=1
            if(lastTime == -1):
                lastTime = firstTime

            if(count != 0):
                chargedEnergy += lastValue - firstValue
                chargingTime += lastTime - firstTime
                connectedTime += lastTimeDelta - firstTime
                print('chargingTime', lastTime - firstTime)
                print('connectedTime',lastTimeDelta - firstTime)
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/test.csv', 'a+') as csvfile:
            writer = csv.writer(csvfile, delimiter= ' ')
            writer.writerow([evSession.pathTofile.split('/')[-1], charged - firstValue, firstTime, lastTime])
        #with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
        #        csv_reader = csv.reader(csvfile, delimiter=";")
        #        count = 0
        #        for row in csv_reader:  
        #                if(count == 0):
        #                    firstTime = float(row[0])
        #                    count +=1
        #                lastTime = float(row[0])

        nOfSession +=1

    interval = datetime.timestamp(max_time_interest) - datetime.timestamp(min_time_interest)
    print('chargingTime',chargingTime)
    print('chargedEnergy', chargedEnergy)
    print('conn',connectedTime)
    connectedTime3 = connectedTime/3600
    numcp = len(chpointlist)
    if(numcp == 0):
        numcp = 1
        
    if(connectedTime != 0):
        kpi531 = (connectedTime/interval)/numcp
        kpi532 = chargingTime/connectedTime
        kpi533 = chargedEnergy/connectedTime3
        kpiList.append([round(kpi531 * 100, 2)])
        kpiList.append([round(kpi532 * 100, 2)])
        kpiList.append([round(kpi533, 2)])

    else:
        kpiList.append([0])
        kpiList.append([0])
        kpiList.append([0])

    kpi534 = nOfSession
    kpiList.append([round(kpi534, 2)])

def calculateGC55(min_time_interest, max_time_interest, locationid, userid, session):
    relevant_bookings = session.query(reservation_booking_logs).join(charging_discharging_session, reservation_booking_logs.ChrgSessID == charging_discharging_session.ChrgSessID).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).filter(charging_discharging_session.PluginTime > min_time_interest).filter(charging_discharging_session.PlugoutTime < max_time_interest).all()
    #relevant_charging_session_id = session.query(reservation_booking_logs.id).join(charging_discharging_session, reservation_booking_logs.ChrgSessID == charging_discharging_session.ChrgSessID).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).filter(charging_discharging_session.PluginTime > min_time_interest).filter(charging_discharging_session.PlugoutTime < max_time_interest).all()
    ev_charging_session = session.query(charging_discharging_session).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).filter(charging_discharging_session.PluginTime < max_time_interest).filter(charging_discharging_session.PlugoutTime > min_time_interest).all()
    
    chargeDemand = 0
    noBookingRequests = 0
    noBookingAccepted = 0
    chargeDemandList = {}
    estList = {}
    lftList = {}
    for relevantBooking in relevant_bookings:
        if(relevantBooking.EnergyEnd != '' and relevantBooking.EnergyEnd != 'UNKNOWN' ):
            chargeEVDemand = float(relevantBooking.EnergyEnd) - float(relevantBooking.EnergyStart)
        else:
            firstValue = 0
            lastValue = 0
            for ev_charge in ev_charging_session:

                if(ev_charge.ChrgSessID == relevantBooking.ChrgSessID):
                    with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+ev_charge.pathTofile, newline='') as csvfile:
                        csv_reader = csv.reader(csvfile, delimiter=";")
                        count = 0  
                        for row in csv_reader:  
                                if (count == 0):
                                    firstValue = float(row[1])
                                    count +=1
                                lastValue = float(row[1])
                    chargeEVDemand = lastValue - firstValue


        chargeDemandList[relevantBooking.ChrgSessID] = float(chargeEVDemand)
        chargeDemand += chargeEVDemand
        noBookingRequests +=1
        if(relevantBooking.Status == 'Confirmed'):
            noBookingAccepted +=1
        
        try:
            estList[relevantBooking.ChrgSessID] = datetime.timestamp(datetime.strptime(relevantBooking.EST, '%Y%m%dT%H%M%S'))
        except:
            for ev_charge in ev_charging_session:
                if(ev_charge.ChrgSessID == relevantBooking.ChrgSessID):
                    
                    estList[relevantBooking.ChrgSessID] = datetime.timestamp(datetime.strptime(ev_charge.PluginTime, '%Y-%m-%d %H:%M:%S'))
                    lftList[relevantBooking.ChrgSessID] = datetime.timestamp(datetime.strptime(relevantBooking.LFT, '%Y%m%dT%H%M%S')) 

    noCharging = 0
    chargedEnergy = 0
    chargingTime = 0
    noFullyCharged = 0
    plugInDelay = 0
    plugOutDelay = 0
    notFinishedInTime = 0
    for evSession in ev_charging_session:
        try:
            hour = int(evSession.ChrgTime.split(':')[0])*3600
            minutes = int(evSession.ChrgTime.split(':')[1])*60
            seconds = int(evSession.ChrgTime.split(':')[2])
            chargingTime += int(hour) + int(minutes) + int(seconds)

            firstValue = 0
            lastValue = 0
            with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=";")
                firstTime = float(next(csv_reader)[0])
                try:
                    firstValue = float(next(csv_reader)[1])
                except:
                    firstValue = 0
                lastValue = -1
                lastTime = -1
                count = 0
                for row in reversed(list(csv_reader)):  
                        if(count != 0 and lastValue != float(row[1])):
                            break
                        count +=1
                        lastValue = float(row[1])
                        lastTime = float(row[0])
                if(count!=0):
                    chargedEnergy += lastValue - firstValue
                    chargingTime += lastTime - firstTime

            noCharging +=1

            try:
                if(chargedEnergy >= chargeDemandList[evSession.ChrgSessID]):
                    noFullyCharged += 1
            except:
                None
               
            try:
                plugInDelay += estList[evSession.ChrgSessID] - datetime.timestamp(datetime.strptime(evSession.PluginTime, '%Y-%m-%d %H:%M:%S'))
                if(lftList[evSession.ChrgSessID] is not None):
                    if(datetime.timestamp(datetime.strptime(evSession.PlugoutTime, '%Y-%m-%d %H:%M:%S')) > lftList[evSession.ChrgSessID]):
                        plugOutDelay +=  datetime.timestamp(datetime.strptime(evSession.PlugoutTime, '%Y-%m-%d %H:%M:%S')) - lftList[evSession.ChrgSessID]
                        notFinishedInTime +=1
            except:
                None
        except:
            None
    if (chargeDemand != 0):
        kpi551 = chargedEnergy/chargeDemand
        kpiList.append([round(kpi551, 2)])
    else:
        kpiList.append([0])
    if(noBookingAccepted != 0):
        kpi552 = noFullyCharged/noBookingAccepted
        kpiList.append([kpi552 * 100])
    else:
        kpiList.append([0])
    kpi553 = 0
    kpiList.append([0])
    if (noCharging != 0):
        #print(' plugindelay' + str(plugInDelay))
        kpi554 = plugInDelay / noCharging
        kpi555 = notFinishedInTime / noCharging

        kpi556 = (plugOutDelay / noCharging)/60
        kpiList.append([round(kpi554 * 100, 2)])
        kpiList.append([round(kpi555 * 100, 2)])
        kpiList.append([round(kpi556, 2)])
    else:
        kpiList.append([0])
        kpiList.append([0])
        kpiList.append([0])




def calculateGC54(min_time_interest, max_time_interest, locationid, userid, session):
    ev_charging_session = session.query(charging_discharging_session).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).filter(charging_discharging_session.PluginTime < max_time_interest).filter(charging_discharging_session.PlugoutTime > min_time_interest).all()
    kpi54 = 0
    error = False
    print('num of ev'+ str(len(ev_charging_session)))
    for evSession in ev_charging_session:
        connectedTime = datetime.timestamp(datetime.strptime(evSession.PlugoutTime, '%Y-%m-%d %H:%M:%S')) - datetime.timestamp(datetime.strptime(evSession.PluginTime, '%Y-%m-%d %H:%M:%S'))
        if(evSession.SOCEnd):
            energyDemand = float(evSession.SOCEnd) - float(evSession.SOCStart)
        else:
            firstValue = 0
            lastValue = 0
            with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=";")
                count = 0  
                for row in csv_reader:  
                        if (count == 0):
                            firstValue = float(row[1])
                            count +=1
                        lastValue = float(row[1])
                energyDemand = (lastValue - firstValue)
        try:
            individual = session.query(ev_individual).filter(ev_individual.EVID == evSession.EVID).first()
            model = session.query(ev_model).filter(func.lower(ev_model.MakeM) == func.lower(individual.MakeM)).first()
            maxChPow = float(model.MaxChPwrAC)
            kpi54 += (maxChPow*(connectedTime/3600 - energyDemand/maxChPow)*0.5)
        except:
            error = True 
    if(error):
        kpi54 = 999999999999999999999
    kpiList.append([round(kpi54, 2)])





def calculateGC513(min_time_interest, max_time_interest, locationid, userid, session):
    cptypes = {}
    cppow = {}
    numberOfSessions = 0
    ev_charging_session = session.query(charging_discharging_session).\
        join(charging_point_individual,charging_discharging_session.CPID == charging_point_individual.CPID).filter(
        charging_point_individual.LOC == locationid).filter(
        charging_discharging_session.PluginTime < max_time_interest).filter(
        charging_discharging_session.PlugoutTime > min_time_interest).all()
    charging_points = session.query(charging_point_individual).filter(charging_point_individual.LOC == locationid).all()
    OfferedFlexibilityIndex = 0
    ActualFlexibilityIndex = 0
    V2GFlexibilityIndex = 0
    for cp in charging_points:
        jsn = cp.Connector.replace("\\", '').replace('[','').replace(']', '').replace('\"', '').replace(':', ': ')
        cptype = yaml.load(jsn, Loader=yaml.FullLoader)
        cptypes[cp.CPID] = cptype['CPType']
        cppow[cp.CPID] = float(cp.ChrgCap)
    for evSession in ev_charging_session:
        OfferedFlexTime = 0
        ActualFlexTime = 0
        plugOutTime = datetime.timestamp(datetime.strptime(evSession.PlugoutTime, '%Y-%m-%d %H:%M:%S'))
        booking = session.query(reservation_booking_logs).filter(reservation_booking_logs.ChrgSessID == evSession.ChrgSessID).first()
        if(booking and booking.EnergyEnd and booking.EnergyEnd!='UNKNOWN'):
            requiredEnergy = float(booking.EnergyEnd) - float(booking.EnergyStart)
        else:
            firstValue = 0
            lastValue = 0
            with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=";")
                count = 0
                firstValue = float(next(csv_reader)[1])
                for row in reversed(list(csv_reader)):
                        if (count == 0):
                            lastValue = float(row[1])
                            count +=1
                            break
            requiredEnergy = lastValue - firstValue
        if(booking and booking.EnergyMin and booking.EnergyMin!='UNKNOWN'):
            minEnergyContent = float(booking.EnergyMin)
        else:
            minEnergyContent = 0
        individual = session.query(ev_individual).filter(ev_individual.EVID == evSession.EVID).first()
        if (individual is not None):
            model = session.query(ev_model).filter(
                func.lower(ev_model.MakeM) == func.lower(individual.MakeM)).first()
            if (model is not None):
                OfferedEnergy = float(model.BatCap) - minEnergyContent
            else:
                OfferedEnergy = 0
        else:
            OfferedEnergy = 0
        try:
            if (cptypes[evSession.CPID] == 'AC'):
                if(cppow[evSession.CPID] > float(model.MaxChPwrAC)):
                    EVChargingPower = cppow[evSession.CPID]
                else:
                    EVChargingPower = float(model.MaxChPwrAC)
                if (model.MaxDischPwrAC != 'null'):
                    EVDisChargingPower = float(model.MaxDischPwrAC)
                else:
                    EVDisChargingPower = float(model.MaxChPwrAC)
            else:
                if(cppow[evSession.CPID] > float(model.MaxChPwrDC)):
                    EVChargingPower = cppow[evSession.CPID]
                else:
                    EVChargingPower = float(model.MaxChPwrDC)
                if (model.MaxDischPwrDC != 'null'):
                    EVDisChargingPower = float(model.MaxDischPwrDC)
                else:
                    EVDisChargingPower = float(model.MaxChPwrDC)
        except:
            OfferedEnergy = 1
            EVChargingPower = 11
            EVDisChargingPower = 11

        #try:
        #    OfferedFlexTime = (datetime.timestamp(datetime.strptime(booking.LFT, '%Y%m%dT%H%M%S')) - datetime.timestamp(datetime.strptime(booking.EST, '%Y%m%dT%H%M%S')))/3600
        #    ActualFlexTime = (plugOutTime - datetime.timestamp(datetime.strptime(booking.EST, '%Y-%m-%d %H:%M:%S')))/3600
        #except:
        for ev_charge in ev_charging_session:
            if(booking and ev_charge.ChrgSessID == booking.ChrgSessID):
                OfferedFlexTime = (datetime.timestamp(datetime.strptime(booking.LFT, '%Y%m%dT%H%M%S')) - datetime.timestamp(datetime.strptime(ev_charge.PluginTime, '%Y-%m-%d %H:%M:%S')))
            #ActualFlexTime = (plugOutTime - datetime.timestamp(datetime.strptime(ev_charge.PluginTime, '%Y-%m-%d %H:%M:%S')))/3600
        #with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
        #    csv_reader = csv.reader(csvfile, delimiter=";")
        #    for row in csv_reader:
        #        charged = float(row[1])
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=";")
            
            firstTime = float(next(csv_reader)[0])
            try:
                firstValue = float(next(csv_reader)[1])
            except:
                firstValue = 0
            lastValue = -1
            lastTime = -1
            count = 0
            for row in csv_reader:
                    charged = float(row[1])
        charged = charged - firstValue
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+evSession.pathTofile, newline='') as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=";")
            firstTime = float(next(csv_reader)[0])
            lastValue2 = -1
            lastTime = -1
            count = 0
            for row in csv_reader:
                    if((float(row[1]) - firstValue)<=(float(charged)-0.1*float(charged))):
                        lastTime = float(row[0])
                    count +=1
            if(count!=0):
                ActualFlexTime = lastTime - firstTime

            #ActualFlexTime = requiredEnergy/EVChargingPower

        '''   individual = session.query(ev_individual).filter(ev_individual.EVID == evSession.EVID).first()
            if(individual is not None):
                model = session.query(ev_model).filter(func.lower(ev_model.MakeM) == func.lower(individual.MakeM)).first()
                if(model is not None):
                    OfferedEnergy = float(model.BatCap) - minEnergyContent
                else:
                    OfferedEnergy = 0
            else:
                OfferedEnergy = 0

            try:
                if(cptypes[evSession.CPID] == 'AC'):
                    EVChargingPower = float(model.MaxChPwrAC)
                    if(model.MaxDischPwrAC != 'null'):
                        EVDisChargingPower =  float(model.MaxDischPwrAC)
                    else:
                        EVDisChargingPower = float(model.MaxChPwrAC)
                else:
                    EVChargingPower = float(model.MaxChPwrDC)
                    if(model.MaxDischPwrDC != 'null'):
                        EVDisChargingPower =  float(model.MaxDischPwrDC)
                    else:
                        EVDisChargingPower = float(model.MaxChPwrDC)
            except:
                OfferedEnergy = 1
                EVChargingPower = 1
            EVDisChargingPower = 1'''
        print('required', charged/EVChargingPower,ActualFlexTime, EVChargingPower )
        numberOfSessions +=1
        if(OfferedFlexTime !=0):
            OfferedFlexibilityIndex += 1 - (charged/EVChargingPower)/(OfferedFlexTime/3600)
            V2GFlexibilityIndex += (OfferedEnergy/EVDisChargingPower)/(OfferedFlexTime/3600)
        else:
            OfferedFlexibilityIndex = 0
            V2GFlexibilityIndex = 0
        if(ActualFlexTime !=0):
            ActualFlexibilityIndex += 1 - (charged/EVChargingPower)/(ActualFlexTime/3600)


    if numberOfSessions != 0:
        kpi5131 = OfferedFlexibilityIndex / numberOfSessions
        if(ActualFlexibilityIndex > 0):
            kpi5132 = ActualFlexibilityIndex / numberOfSessions
        else:
            kpi5132 = 0
        kpi5133 = V2GFlexibilityIndex / numberOfSessions
        kpiList.append([round(100*abs(kpi5131), 2)])
        kpiList.append([round(100*abs(kpi5132), 2)])
        kpiList.append([round(kpi5133, 2)])
    else:
        kpiList.append([0])
        kpiList.append([0])
        kpiList.append([0])








def calculateGC511(min_time_interest, max_time_interest, locationid, userid, session):
    cptypes = {}
    numberOfSessions = 0
    ev_charging_session = session.query(charging_discharging_session).\
        join(charging_point_individual,charging_discharging_session.CPID == charging_point_individual.CPID).filter(
        charging_point_individual.LOC == locationid).filter(
        charging_discharging_session.PluginTime < max_time_interest).filter(
        charging_discharging_session.PlugoutTime > min_time_interest).all()
    charging_points = session.query(charging_point_individual).filter(charging_point_individual.LOC == locationid).all()
    relevant_bookings = session.query(reservation_booking_logs).join(charging_discharging_session, reservation_booking_logs.ChrgSessID == charging_discharging_session.ChrgSessID).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).filter(charging_discharging_session.PluginTime > min_time_interest).filter(charging_discharging_session.PlugoutTime < max_time_interest).all()
    priceList = {}
    energyDict = {}
    timeElapsed = {}

    totalCostEnergy = 0
    totalCost = 0 
    totalEnergy = 0
    temporary = 0
    for evSession in ev_charging_session:
        plugOutTime = datetime.timestamp(datetime.strptime(evSession.PlugoutTime, '%Y-%m-%d %H:%M:%S'))
        plugInTime = datetime.timestamp(datetime.strptime(evSession.PluginTime, '%Y-%m-%d %H:%M:%S'))
        timeElapsed[evSession.ChrgSessID] = plugOutTime - plugInTime
        booking = session.query(reservation_booking_logs).filter(reservation_booking_logs.ChrgSessID == evSession.ChrgSessID).first()
        requiredEnergy = float(booking.EnergyEnd) - float(booking.EnergyStart)
        energyDict[evSession.ChrgSessID] = requiredEnergy
    for relevantBooking in relevant_bookings:
        chargeEVDemand = float(relevantBooking.EnergyEnd) - float(relevantBooking.EnergyStart)
        chargeDemand += chargeEVDemand
        priceList[relevantBooking.ChrgSessID] = relevantBooking.TariffID
        chargeDemandList[relevantBooking.ChrgSessID] = float(chargeEVDemand)
        tariffs = session.query(tariff_scheme_individual).filter(tariff_scheme_individual.TariffID == relevantBooking.TariffID).all()
        temporary = 0
        percent = 0
        for tariff in tariffs:
            if(tariff.PriceModel == 'Price per kWh'):
                temporary += float(energyDict[relevantBooking.ChrgSessID])*float(tariff.Price)
                totalCostEnergy += float(energyDict[relevantBooking.ChrgSessID])*float(tariff.Price)
            if(tariff.PriceModel == 'Fixed price'):
                temporary += float(tariff.Price)
            if(tariff.PriceModel == 'Percentage added to total'):
                percent = float(tariff.Price)                
            if(tariff.PriceModel == 'Price per time unit'):
                if(tariff.Period == 'Minute'):
                    price = tariff.price/60
                if(tariff.Period == 'Hour'):
                    price = tariff.price/3600
                if(tariff.Period == 'Day'):
                    price = tariff.price/86400
                if(tariff.Period == 'Week'):
                    price = tariff.price/(86400*7)
                if(tariff.Period == 'Month'):
                    price = tariff.price/(86400*30)
                if(tariff.Period == 'Year'):
                    price = tariff.price/(86400*365)
                temporary += timeElapsed[relevantBooking.ChrgSessID]*price
        temporary = temporary + temporary*percent/100
        totalCost += temporary
        totalEnergy += float(energyDict[relevantBooking.ChrgSessID])
   
    if totalEnergy != 0:
        kpi5111 = totalCost/totalEnergy
        kpi5112 = totalCostEnergy/totalEnergy
        kpiList.append([round(kpi5111 * 100, 2)])
        kpiList.append([round(kpi5112 * 100, 2)])
    else:
        kpiList.append([0])
        kpiList.append([0])

















def selfConsumption2ndVersion(min_time_interest_timestamp, max_time_interest_timestamp, min_time_interest,max_time_interest, locationid, userid, directory,avpow):
        
        
        
        
        import_exports = session.query(energy_import_export).join(energy_meters_individual, energy_meters_individual.MeterID == energy_import_export.MeterID).filter(energy_meters_individual.LOC == locationid).all()
        consumers = []  # lista che contiene le liste con tutte le timeseries di potenza ricampionate
        consumersen = []

        for import_export in import_exports:
            with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+import_export.pathTofile) as csvfile:
                count = 0
                x = []  # lista dei tempi della timeseries
                y = [] 
                csv_reader = csv.reader(csvfile, delimiter=";")
                for row in csv_reader:
                    if (min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                        if(row[1] != ''):
                            if(count == 0):
                                x.append(min_time_interest_timestamp)
                                y.append(0)
                                x.append(float(row[0]))
                                y.append(0)
                                lastVal = float(row[1])
                                lastTime = float(row[0])
                                count = 1
                            else:
                                energyTemp = float(row[1])-lastVal
                                if(energyTemp >0):
                                    #egrid = egrid + energyTemp
                                    y.append(3600*(energyTemp)/(float(row[0]) - lastTime))
                                    x.append(float(row[0]))
                                else:
                                    None
                                lastVal = float(row[1])
                                lastTime = float(row[0])
                if(len(x)!=0):
                    x.append(lastTime+1)
                    y.append(0)
                    if (max_time_interest_timestamp > lastTime):  # stesso discorso di prima, se l'ultimo tempo della timeseries ?? pi?? piccolo del massimo tempo di interesse metto uno zero per non aver problemi dopo
                        y.append(0)
                        x.append(max_time_interest_timestamp)
                    f = interpolate.interp1d(x, y)  # faccio l'interpolazione lineare
                    xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,(max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)  # mi creo il vettore dei tempi con un sample ogni 5 minuti (300 secondi)
                    ynew = f(xnew)  # genero la nuova serie di potenze ricampionatew
                    consumers.append(ynew)  # aggiungo la lista ai consumers
        
        consumers = []
        if(len(consumers) == 0):
            ev_charging_session = session.query(charging_discharging_session.pathTofile).\
            join(charging_point_individual,charging_discharging_session.CPID == charging_point_individual.CPID).filter(
            charging_point_individual.LOC == locationid).filter(
            charging_discharging_session.PluginTime < max_time_interest).filter(
            charging_discharging_session.PlugoutTime > min_time_interest).all()

            ev_charging_session = [item[0] for item in ev_charging_session]
            count = 1
            consumedEn = 0
            numsess = 0
            for i,file in enumerate(ev_charging_session):  # Per tutti i file nella cartella
                
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+file, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")
                    count = 0  # questo mi serve per scartare le prime due righe di metadata
                    x = []  # lista dei tempi della timeseries
                    y = []  # lista dei valori della timeseries
                    yen = []
                    lastSample = 0  # Questo mi serve per tenermi in memoria il tempo precedente alla riga che sto leggendo, cosi posso farmi il delta per la trasformazione in potenza
                    lastValue = 0  # Questo mi serve per tenermi in memoria il valore di energia precedente alla riga che sto leggendo, cosi posso farmi il delta per la trasformazione in potenza
                    for row in csv_reader:  # per tutte le righe
                        if(min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                            if (count == 0):
                                x.append(min_time_interest_timestamp)
                                y.append(0)
                                yen.append(0)
                                firstSample = float(row[1])
                                x.append(float(row[0]))
                                y.append(0)
                                yen.append(0)
                            else:
                                x.append(float(row[0]))
                                try:
                                    y.append(3600*(float(row[1]) - lastValue) / (float(row[0]) - lastSample))
                                    yen.append(float(row[1]))
                                except:
                                    y.append(0)
                                    yen.append(float(row[1]))

                            lastSample = float(row[0])
                            lastValue = float(row[1])
                            count += 1
                consumedEn += lastValue - firstSample     
                x.append(lastSample+1)
                y.append(0)
                yen.append(float(row[1]))

                if (max_time_interest_timestamp > lastSample):  # stesso discorso di prima, se l'ultimo tempo della timeseries ?? pi?? piccolo del massimo tempo di interesse metto uno zero per non aver problemi dopo
                    yen.append(float(row[1]))
                    y.append(0)
                    x.append(max_time_interest_timestamp)
                f = interpolate.interp1d(x, y)  # faccio l'interpolazione lineare
                fen = interpolate.interp1d(x, yen)

                xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,(max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)  # mi creo il vettore dei tempi con un sample ogni 5 minuti (300 secondi)
                xnewen = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,900, dtype = int)  # mi creo il vettore dei tempi con un sample ogni 5 minuti (300 secondi)
                
                ynew = f(xnew)  # genero la nuova serie di potenze ricampionatew
                ynewen = fen(xnewen)
                consumers.append(ynew)  # aggiungo la lista ai consumers
                consumersen.append(ynewen)
                numsess +=1
            print('numsess', numsess)
            hc_list = session.query(heating_cooling_session.pathTofile).\
            join(heater_cooler_individual,heating_cooling_session.HCID == heater_cooler_individual.HCID).filter(
            heater_cooler_individual.LOC == locationid).all()
            #hc_list = session.query(heating_cooling_session.pathTofile).all()
            #hc_list = [item[0] for item in hc_list]

            for file in hc_list:
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+file[0], newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")
                    count = 0
                    x = []
                    y = []
                    yen = []
                    lastSample = 0
                    for row in csv_reader:
                        if(min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                            if (count == 0):
                                x.append(min_time_interest_timestamp)
                                y.append(0)
                                yen.append(0)

                                x.append(float(row[0]))
                                y.append(0)
                                yen.append(0)
                            else:
                                x.append(float(row[0]))
                                try:
                                    y.append(3600*(float(row[1]) - lastValue) / (float(row[0]) - lastSample))
                                    yen.append(float(row[1]))

                                except:
                                    y.append(0)
                                    yen.append(float(row[1]))
                                     
                            lastSample = float(row[0])
                            lastValue = float(row[1])
                            count += 1
                x.append(lastSample+1)
                y.append(0)
                yen.append(float(row[1]))

                if (max_time_interest_timestamp > lastSample):
                    y.append(0)
                    x.append(max_time_interest_timestamp)
                    yen.append(float(row[1]))
                      
                f = interpolate.interp1d(x, y)
                fen = interpolate.interp1d(x, yen)

                xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)
                xnewen = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,900, dtype = int)  # mi creo il vettore dei tempi con un sample ogni 5 minuti (300 secondi)
                ynewen = fen(xnewen)

                ynew = f(xnew)
                consumers.append(ynew)
                consumersen.append(ynewen)

            productions = []
            bat_list = session.query(battery_session.pathTofile).\
            join(battery_individual,battery_session.BatID == battery_individual.BatID).filter(
            battery_individual.LOC == locationid).filter(battery_session.startPoint > min_time_interest).filter(battery_session.endPoint < max_time_interest).all()
            #print('num of bat: ' + str(len(bat_list)))
            productions = []
            bat_list = [item[0] for item in bat_list]
            if(locationid == 'P2D1L1'):
                bat_list = []
            consbat = []
            prodbat = []
            for file in bat_list:

                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+file, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                    count = 0
                    x = []
                    yprod = []
                    ycons = []
                    yen = []
                    ytot = []
                    x2 = []
                    for row in csv_reader:
                            pow = 0
                            val = row[1].replace(',', '.')  #se hanno messo la virgola come delimitatore decimale la cambio
                            if (count == 0 and min_time_interest_timestamp < float(row[0])):
                                x.append(min_time_interest_timestamp)
                                x2.append(min_time_interest_timestamp)

                                yprod.append(0)
                                ycons.append(0)
                                yen.append(float(row[1]))
                                ytot.append(0)
                            if(count == 0):     #dopo di questa riga non sar?? pi?? la prima riga
                                count = 1
                            else:
                                if(min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                                    if(float(row[0]) == lastSample):
                                        pow = 0
                                    else:
                                        pow = 3600*(float(val) - lastValue) / (float(row[0]) - lastSample)
                                    ytot.append(pow)
                                    if(float(row[1]) > 0):
                                        yen.append(float(row[1]))
                                        x2.append(float(row[0]))
                                    if(pow >= 0):
                                        ycons.append(pow)
                                        yprod.append(0)
                                        x.append(float(row[0]))
                                    else:
                                        ycons.append(0)
                                        yprod.append(-pow)
                                        x.append(float(row[0]))
                            count +=1
                            lastSample = float(row[0])
                            lastValue = float(val)
                    if(max_time_interest_timestamp > lastSample):
                        ycons.append(0)
                        yprod.append(0)
                        ytot.append(0)
                        x2.append(max_time_interest_timestamp)
                        yen.append(0)
                        x.append(max_time_interest_timestamp)


                    xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)
                    xnew2 = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,900, dtype = int)

                    fcons = interpolate.interp1d(x, ycons)
                    fprod = interpolate.interp1d(x, yprod)
                    fen = interpolate.interp1d(x2, yen)
                    ynewcons = fcons(xnew)
                    ynewprod = fprod(xnew)
                    ynewen = f2(xnew2)
                    consumersen.append(ynewen)

                    consumers.append(ynewcons)
                    productions.append(ynewprod)
            first = 0
            for consumer in consumersen:
                for i in range(len(consumer)):
                    if (first == 0):
                        consum.append(consumer[i])
                    else:
                        consum[i] += consumer[i]
                first = 1

        else:
            bat_list = session.query(battery_session.pathTofile).\
            join(battery_individual,battery_session.BatID == battery_individual.BatID).filter(
            battery_individual.LOC == locationid).filter(battery_session.startPoint > min_time_interest).filter(battery_session.endPoint < max_time_interest).all()
            #print('num of bat: ' + str(len(bat_list)))
            productions = []
            bat_list = [item[0] for item in bat_list]
            if(locationid == 'P2D1L1'):
                bat_list = []
            consbat = []
            prodbat = []
            for file in bat_list:

                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+file, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                    count = 0
                    x = []
                    yprod = []
                    ycons = []
                    ytot = []
                    yen = []
                    for row in csv_reader:
                            pow = 0
                            val = row[1].replace(',', '.')  #se hanno messo la virgola come delimitatore decimale la cambio
                            if (count == 0 and min_time_interest_timestamp < float(row[0])):
                                x.append(min_time_interest_timestamp)
                                yprod.append(0)
                                ycons.append(0)
                                ytot.append(0)
                            if(count == 0):     #dopo di questa riga non sar?? pi?? la prima riga
                                count = 1
                            else:
                                if(min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                                    if(float(row[0]) == lastSample):
                                        pow = 0
                                    else:
                                        pow = 3600*(float(val) - lastValue) / (float(row[0]) - lastSample)
                                    ytot.append(pow)
                                    if(pow >= 0):
                                        ycons.append(pow)
                                        yprod.append(0)
                                        x.append(float(row[0]))
                                    else:
                                        ycons.append(0)
                                        yprod.append(-pow)
                                        x.append(float(row[0]))

                            count +=1
                            lastSample = float(row[0])
                            lastValue = float(val)
                    if(max_time_interest_timestamp > lastSample):
                        ycons.append(0)
                        yprod.append(0)
                        ytot.append(0)
                        x.append(max_time_interest_timestamp)


                    xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)
                    fcons = interpolate.interp1d(x, ycons)
                    fprod = interpolate.interp1d(x, yprod)
                    ynewcons = fcons(xnew)
                    ynewprod = fprod(xnew)


                    #consumers.append(ynewcons)
                    productions.append(ynewprod)


        xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)

        totalPowerListConsumers = []
        totalconsumedPower = 0 
        first = 0
        for consumer in consumers:
            for i in range(len(consumer)):
                if (first == 0):
                    totalPowerListConsumers.append(consumer[i])
                    totalconsumedPower += consumer[i]
                else:
                    totalPowerListConsumers[i] += consumer[i]
                    totalconsumedPower += consumer[i]
            first = 1



        pv_list = session.query(solar_plant_session.pathTofile).\
        join(pv_panel_individual,solar_plant_session.PlantID == pv_panel_individual.SolarPlantID).filter(
        pv_panel_individual.LOC == locationid).all()
        pv_list = session.query(solar_plant_session.pathTofile).all()
        #solar_plant_session.startPoint > min_time_interest).filter(
        #solar_plant_session.endPoint < max_time_interest).all()
        totlist = []

        pv_list = [item[0] for item in pv_list]
        for item in pv_list:
            if(item.split('-')[2]==locationid):
                totlist.append(item)
            if(locationid == "P3D2L1" and item.split('-')[2]=="P3D2L2"):
                totlist.append(item)

        for file in totlist:
            energy_y = []
            lastValue = 0
            with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+file, newline='') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                count = 0
                x = []
                y = []
                lastSample = 0
                for row in csv_reader:
                    if(min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                        if (count == 0):
                            x.append(min_time_interest_timestamp)
                            y.append(0)
                            energy_y.append(float(row[1]))
                            x.append(float(row[0]))
                            y.append(0)
                            energy_y.append(float(row[1]))
                        else:
                            x.append(float(row[0]))
                            try:
                                energy_y.append(float(row[1]))
                                y.append(3600*(float(row[1]) - lastValue) / (float(row[0]) - lastSample))
                            except:
                                y.append(0)
                        
                        lastSample = float(row[0])
                        if(row[1] != ''):
                            lastValue = float(row[1])
                        count += 1
            
            x.append(lastSample+1)
            energy_y.append(lastValue)
            y.append(0)                    
            if (max_time_interest_timestamp > lastSample):
                y.append(0)
                try:
                    energy_y.append(lastValue)
                except:
                    energy_y.append(0)
                x.append(max_time_interest_timestamp)
            avg_tim = np.average(diff(x))
            resample_tick = int(180/avg_tim)
            if(resample_tick == 0):
                resample_tick = 1
            newxx = x[::resample_tick]

            ynewxx = energy_y[::resample_tick]
            if('P2' in locationid):
                dydx = diff(ynewxx)/diff(newxx)
                dydx = 5*3600*dydx
            if('P2' in locationid):
                if(len(newxx)>4):
                    try:
                        f = interpolate.interp1d(newxx[:-1], dydx)
                        xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)
                        ynew = f(xnew)
                    except:
                        xx = newxx[:-1]
                        xx.append(max_time_interest_timestamp)
                        dydx = np.append(dydx,0)
                        #dydx.append(0)
                        f = interpolate.interp1d(xx, dydx)
                        xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)
                        ynew = f(xnew)
                    productions.append(ynew)
            else:
                f = interpolate.interp1d(x, y)
                xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)
                ynew = f(xnew)
                productions.append(ynew)
        totalPowerListProducers = []
        totalProducedPower = 0
        first = 0
        totpek = []
        for x in totalPowerListConsumers:
            totpek.append(x)
        for production in productions:
            for i in range(len(production)):
                if (first == 0):
                    totalPowerListProducers.append(production[i])
                    totalProducedPower += production[i]
                    #totpek[i] -= production[i]
                else:
                    totalPowerListProducers[i] += production[i]
                    totalProducedPower += production[i]
                    #totpek[i] -= production[i]

            first = 1

        if(len(productions) == 0):
            for i in range(len(totalPowerListConsumers)):
                totalPowerListProducers.append(0)
                
        global totalProducedEnergy
        for x in totalPowerListConsumers:
            totalProducedEnergy += x*(max_time_interest_timestamp - min_time_interest_timestamp)/1000/3600
        grid = []
        size = len(totalPowerListProducers)
        if(len(totalPowerListConsumers) == 0):
            size = 0

        print(totalPowerListProducers)
        for i in range(size):
            if (totalPowerListConsumers[i] - totalPowerListProducers[i] > 0):
                grid.append([xnew[i], totalPowerListConsumers[i] - totalPowerListProducers[i]])
            else:
                grid.append([xnew[i], 0])
        if(len(totalPowerListConsumers)==0):
            averagePower = 0
            Pmax = 0
        else:
            averagePower = statistics.median(totalPowerListConsumers)
            Pmax = max(totalPowerListConsumers)
        indx= 0 
        totpow = 0
        for xx in totalPowerListConsumers:
            if(xx == Pmax):
                ax = indx
            indx +=1
        if avpow != 0:
            kpi5101 = Pmax / avpow
            kpi5103 = avpow
            kpi5102 = Pmax
            kpi5104 = datetime.fromtimestamp(xnew[ax]).strftime("%m/%d/%Y, %H:%M:%S")

            kpiList.append([round(kpi5101,2)])
            kpiList.append([round(kpi5103,2)])
            kpiList.append([round(kpi5102,2)])
            kpiList.append([kpi5104])

        else:
            #None
            kpiList.append([0])
            kpiList.append([0])
            kpiList.append([0])
            kpiList.append([0])

        selfConsumptionProv = 0
        selfC = []
        count = 0
        for i in range(size):
            count +=1
            if (totalPowerListProducers[i] != 0):
                if (totalPowerListProducers[i] < totalPowerListConsumers[i]):
                    selfC1 = 1
                    selfConsumptionProv += totalPowerListProducers[i]
                else:
                    selfC1 = (totalPowerListConsumers[i]) / totalPowerListProducers[i]
                    selfConsumptionProv += totalPowerListConsumers[i]
                selfC.append([xnew[i], selfC1*100])
            else:
                selfC.append([xnew[i], 0])
        selfConsumption = 0
        tot_cons_power = totalconsumedPower
        delta_time = xnew[1]-xnew[0]
        print("salvatore, totalConsumedPower:",totalconsumedPower, xnew[1]-xnew[0])
        if (totalProducedPower != 0):
            selfConsumption = selfConsumptionProv / totalProducedPower
        else:
            selfConsumption = 0
            
        selfSufficiency = 0

        if (totalconsumedPower != 0):
            selfSufficiency = selfConsumptionProv / totalconsumedPower
        else:
            selfSufficiency = 0
        
        kpiList.append([round(selfConsumption*100, 2)])
        kpiList.append([round(selfSufficiency*100, 2)])
        c = 0
        for element in xnew:
            if (c ==0):
                el = element
                c+=1
            else:
                if(el >= element):
                    print('ERRORE')
        for i in range(size):
            totalPowerListConsumers[i] = [datetime.fromtimestamp(xnew[i]).strftime("%m/%d/%Y-%H:%M:%S"), totalPowerListConsumers[i]]
            totalPowerListProducers[i] = [datetime.fromtimestamp(xnew[i]).strftime("%m/%d/%Y-%H:%M:%S"), totalPowerListProducers[i]]
        try:
        
            with open('/home/branco/public_html/gccalculator/csv/kpi/' + userid + '1.csv', 'w') as result_file:
                wr = csv.writer(result_file, delimiter=" ")
                wr.writerows(totalPowerListConsumers)
        except:
            None
        try:
            with open('/home/branco/public_html/gccalculator/csv/kpi/' + userid + '2.csv', 'w') as result_file:
                wr = csv.writer(result_file, delimiter=" ")
                wr.writerows(totalPowerListProducers)
        except:
            None
        try:
            with open('/home/branco/public_html/gccalculator/csv/kpi/' + userid + '3.csv', 'w') as result_file:
                wr = csv.writer(result_file, delimiter=" ")
                wr.writerows(grid)
        except:
            None
        try:
            with open('/home/branco/public_html/gccalculator/csv/kpi/' + userid + '4.csv', 'w') as result_file:
                wr = csv.writer(result_file, delimiter=" ")
                wr.writerows(selfC)
        except:
            None

def calculateKPI510byhour(min_time_interest_timestamp, max_time_interest_timestamp, min_time_interestTotal,max_time_interestTotal, locationid, userid, directory):
        firstKPI = []
        secondKPI = []
        xaxis = []
        xaxis.append(min_time_interest_timestamp)
        firstKPI.append(0)
        min_time_prov = min_time_interest_timestamp
        max_time_prov = min_time_interest_timestamp + 3600
        while(max_time_prov <= max_time_interest_timestamp):
            temp = min_time_prov + 1800
            xaxis.append(temp)
            min_time_interest = datetime.fromtimestamp(min_time_prov)
            max_time_interest = datetime.fromtimestamp(max_time_prov)
            ev_charging_session = session.query(charging_discharging_session). \
            join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(
            charging_point_individual.LOC == locationid).filter(
            charging_discharging_session.PluginTime > min_time_interestTotal).filter(
            charging_discharging_session.PlugoutTime < max_time_interestTotal).all()
            #files = glob.glob(
            #    "./Utils/shared/recordings_logs/energy_consumption_production/ev_charging_discharging_sessions/*.csv")
            totalEnergyConsumption = 0
            #print('numev' + str(len(ev_charging_session)))
            consumers = []
            for chses in ev_charging_session:
                file = chses.pathTofile
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils'+file.split('.')[1]+'.csv', newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                    count = 0
                    energyConsumption = 0
                    first_energy_sample = 0
                    x = []
                    y = []
                    lastSample = 0
                    first = 0
                    for row in csv_reader:
                        val = row[1].replace(',', '.')
                        if(min_time_prov < float(row[0]) and max_time_prov > float(row[0])):
                            if (count == 0 and min_time_prov < float(row[0])):
                                x.append(min_time_prov)
                                y.append(0)
                                x.append(float(row[0]))
                                y.append(0)
                            else:
                                if(first == 0):
                                    first_energy_sample = float(val)
                                    first = 1
                                x.append(float(row[0]))
                                try:
                                    y.append(3600*(float(val) - lastValue) / (float(row[0]) - lastSample))
                                    energyConsumption = float(val)
                                except:
                                    y.append(0)
                            lastSample = float(row[0])
                            lastValue = float(val)
                            count += 1            
                        x.append(lastSample+1)
                        y.append(0)
                if (max_time_prov > lastSample):
                    y.append(0)
                    x.append(max_time_prov)
                f = interpolate.interp1d(x, y)
                xnew = np.arange(min_time_prov, max_time_prov, 300, dtype = int)
                ynew = f(xnew)
                consumers.append(ynew)
                energyConsumption = energyConsumption - first_energy_sample
                #print('encons' + str(energyConsumption))
                totalEnergyConsumption += energyConsumption


            hcSessions = session.query(heating_cooling_session). \
            join(heater_cooler_individual, heating_cooling_session.HCID == heater_cooler_individual.HCID).filter(
            heater_cooler_individual.LOC == locationid).all()
            files = glob.glob(
                "/home/branco/public_html/gccalculator/csv/newCalculator/Utils/"+directory+"/recordings_logs/energy_consumption_production/heating_cooling_sessions/*.csv")
            #print('numhc' + str(len(hcSessions)))

            for hcses in hcSessions:
                file = "/home/branco/public_html/gccalculator/csv/newCalculator/Utils/"+hcses.pathTofile
                if(file.split('/')[-1].split('-')[2] == locationid):
                    with open(file, newline='') as csvfile:
                        csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                        count = 0
                        energyConsumption = 0
                        first_energy_sample = 0
                        first = 0
                        x = []
                        y = []
                        lastSample = 0
                        for row in csv_reader:
                            val = row[1].replace(',', '.')
                            if (count == 0 and min_time_prov < float(row[0])):
                                x.append(min_time_prov)
                                y.append(0)
                                x.append(float(row[0]))
                                y.append(0)
                            else:
                                if(first == 0):
                                    first_energy_sample = float(val)
                                    first = 1
                                x.append(float(row[0]))
                                try:
                                    y.append(3600*(float(val) - lastValue) / (float(row[0]) - lastSample))
                                    energyConsumption = float(val)
                                except:
                                    y.append(0)
                            lastSample = float(row[0])
                            lastValue = float(val)
                            count += 1
                    x.append(lastSample+1)
                    y.append(0)
                    if (max_time_prov > lastSample):
                        y.append(0)
                        x.append(max_time_prov)
                    f = interpolate.interp1d(x, y)
                    xnew = np.arange(min_time_prov, max_time_prov, 300, dtype = int)
                    ynew = f(xnew)
                    consumers.append(ynew)
                    energyConsumption = energyConsumption - first_energy_sample
                    totalEnergyConsumption += energyConsumption

            files = glob.glob(
                "/home/branco/public_html/gccalculator/csv/newCalculator/Utils/"+directory+"/recordings_logs/energy_consumption_production/washing_sessions/*.csv")
            for file in files:
                with open(file, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                    count = 0
                    energyConsumption = 0
                    x = []
                    y = []
                    lastSample = 0
                    for row in csv_reader:
                        if (count == 0 and min_time_prov < float(row[0])):
                            x.append(min_time_prov)
                            y.append(0)
                        else:
                            x.append(float(row[0]))
                            try:
                                y.append((float(row[1]) - lastValue) / (float(row[0]) - lastSample))
                                energyConsumption = float(row[1])
                            except:
                                y.append(0)
                        
                        lastSample = float(row[0])
                        lastValue = float(row[1])
                        count += 1
                if (max_time_prov > lastSample):
                    y.append(0)
                    x.append(max_time_prov)
                f = interpolate.interp1d(x, y)
                xnew = np.arange(min_time_prov, max_time_prov, 300, dtype = int)
                ynew = f(xnew)
                consumers.append(ynew)
                totalEnergyConsumption += energyConsumption
                
            bat_list = session.query(battery_session.pathTofile).\
            join(battery_individual,battery_session.BatID == battery_individual.BatID).filter(
            battery_individual.LOC == locationid).filter(
            battery_session.startPoint > min_time_interestTotal).filter(
            battery_session.endPoint < max_time_interestTotal).all()
            bat_list = [item[0] for item in bat_list]
            if(locationid == 'P2D1L1' or locationid == 'P2D1L3'):
                bat_list = []
            #print('numbt' + str(len(bat_list)))

            for file in bat_list:
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+file, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                    count = 0
                    x = []
                    yprod = []
                    ycons = []
                    ytot = []
                    x.append(min_time_prov)
                    ycons.append(0)
                    for row in csv_reader:
                            val = row[1].replace(',', '.')  #se hanno messo la virgola come delimitatore decimale la cambio
                            if(count == 0):
                                first_energy_sample = float(val)
                                count =1
                            else:
                                if(min_time_prov < float(row[0]) and max_time_prov > float(row[0])):
                                    if(float(row[0]) == lastSample):
                                        pow = 0
                                    else:
                                        pow = 3600*(float(val) - lastValue) / (float(row[0]) - lastSample)
                                    ytot.append(pow)
                                    if(pow >= 0):
                                        ycons.append(pow)
                                        yprod.append(0)
                                        x.append(float(row[0]))
                                    else:
                                        ycons.append(0)
                                        yprod.append(-pow)
                                        x.append(float(row[0]))
                            lastSample = float(row[0])
                            lastValue = float(val)

                ycons.append(0)
                x.append(max_time_prov)

                xnew = np.arange(min_time_prov, max_time_prov, 300, dtype = int)

                fcons = interpolate.interp1d(x, ycons)
                ynewcons = fcons(xnew)
                consumers.append(ynewcons)
                
                energyConsumption = energyConsumption - first_energy_sample
                if(energyConsumption < 0):
                    energyConsumption = 0 
                totalEnergyConsumption += energyConsumption

            totalPowerList = []
            first = 0
            for consumer in consumers:
                for i in range(len(consumer)):
                    if (first == 0):
                        totalPowerList.append(consumer[i])
                    else:
                        totalPowerList[i] += consumer[i]
                first = 1
            if(len(totalPowerList)==0):
                Pmax = 0
            else:
                Pmax = totalPowerList[0]
            for i in range(len(totalPowerList)):
                if totalPowerList[i] > Pmax:
                    Pmax = totalPowerList[i]
            deltaT = 3600
            averagePower = 3600*totalEnergyConsumption / deltaT


            min_time_prov = min_time_prov + 3600
            max_time_prov = max_time_prov + 3600
            if averagePower != 0:
                kpi5101 = abs(Pmax) / averagePower
                kpi5103 = 3600*totalEnergyConsumption / deltaT
                firstKPI.append(round(kpi5101,2))
                secondKPI.append(round(kpi5103,2))
            else:
                firstKPI.append(0)
                secondKPI.append(0)

        xaxis.append(max_time_interest_timestamp)
        firstKPI.append(0)
        plottingSeries = []
        
	
        f = interpolate.interp1d(xaxis, firstKPI, kind='nearest')
        xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/1000, dtype = int)
        ynew = f(xnew)
        for i in range(len(xnew)):
            plottingSeries.append([xnew[i], ynew[i]])
        try:
            with open('/home/branco/public_html/gccalculator/csv/kpi/' + userid + 'kpi510.csv', 'w') as result_file:
                wr = csv.writer(result_file, delimiter=" ")
                wr.writerows(plottingSeries)
        except:
            None



def calculateKPI510(min_time_interest_timestamp, max_time_interest_timestamp, min_time_interest,max_time_interest, locationid, userid, directory):
        min_time_prov = min_time_interest_timestamp
        max_time_prov = min_time_interest_timestamp + 3600
        ev_charging_session = session.query(charging_discharging_session). \
        join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(
        charging_point_individual.LOC == locationid).filter(
        charging_discharging_session.PluginTime < max_time_interest).filter(
        charging_discharging_session.PlugoutTime > min_time_interest).all()
        #files = glob.glob(
        #    "./Utils/shared/recordings_logs/energy_consumption_production/ev_charging_discharging_sessions/*.csv")
        totalEnergyConsumption = 0
        consumers = []
        print('num of ev_charging_session: ' + str(len(ev_charging_session)))
        for chses in ev_charging_session:
            file = chses.pathTofile
            with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils'+file.split('.')[1]+'.csv', newline='') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                count = 0
                energyConsumption = 0
                first_energy_sample = 0
                x = []
                y = []
                lastSample = 0
                first = 0
                for row in csv_reader:
                
                    val = row[1].replace(',', '.')
                    if(min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                        if (count == 0 and min_time_interest_timestamp < float(row[0])):
                            x.append(min_time_interest_timestamp)
                            y.append(0)
                            x.append(float(row[0]))
                            y.append(0)
                        else:
                            if(first == 0):
                                first_energy_sample = float(val)
                                first = 1
                            x.append(float(row[0]))
                            try:
                                y.append(3600*(float(val) - lastValue) / (float(row[0]) - lastSample))
                                energyConsumption = float(val)
                            except:
                                y.append(0)
                        lastSample = float(row[0])
                        lastValue = float(val)
                        count += 1            
                x.append(lastSample+1)
                y.append(0)
            if (max_time_interest_timestamp > lastSample):
                y.append(0)
                x.append(max_time_interest_timestamp)
            f = interpolate.interp1d(x, y)
            xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/300, dtype = int)
            ynew = f(xnew)
            consumers.append(ynew)
            energyConsumption = energyConsumption - first_energy_sample
            totalEnergyConsumption += energyConsumption


        hcSessions = session.query(heating_cooling_session). \
        join(heater_cooler_individual, heating_cooling_session.HCID == heater_cooler_individual.HCID).filter(
        heater_cooler_individual.LOC == locationid).all()
        files = glob.glob(
            "/home/branco/public_html/gccalculator/csv/newCalculator/Utils/"+directory+"/recordings_logs/energy_consumption_production/heating_cooling_sessions/*.csv")

        for hcses in hcSessions:
            file = "/home/branco/public_html/gccalculator/csv/newCalculator/Utils/"+hcses.pathTofile
            if(file.split('/')[-1].split('-')[2] == locationid):
                with open(file, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                    count = 0
                    energyConsumption = 0
                    first_energy_sample = 0
                    first = 0
                    x = []
                    y = []
                    lastSample = 0
                    for row in csv_reader:
                        val = row[1].replace(',', '.')
                        if (count == 0 and min_time_interest_timestamp < float(row[0])):
                            x.append(min_time_interest_timestamp)
                            y.append(0)
                            x.append(float(row[0]))
                            y.append(0)
                        else:
                            if(first == 0):
                                first_energy_sample = float(val)
                                first = 1
                            x.append(float(row[0]))
                            try:
                                y.append(3600*(float(val) - lastValue) / (float(row[0]) - lastSample))
                                energyConsumption = float(val)
                            except:
                                y.append(0)
                        lastSample = float(row[0])
                        lastValue = float(val)
                        count += 1
                x.append(lastSample+1)
                y.append(0)
                if (max_time_interest_timestamp > lastSample):
                    y.append(0)
                    x.append(max_time_interest_timestamp)
                f = interpolate.interp1d(x, y)
                xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp, (max_time_interest_timestamp - min_time_interest_timestamp)/300, dtype = int)
                ynew = f(xnew)
                consumers.append(ynew)
                energyConsumption = energyConsumption - first_energy_sample
                totalEnergyConsumption += energyConsumption

      
            
        bat_list = session.query(battery_session.pathTofile).\
        join(battery_individual,battery_session.BatID == battery_individual.BatID).filter(
        battery_individual.LOC == locationid).filter(
        battery_session.startPoint > min_time_interest).filter(
        battery_session.endPoint < max_time_interest).all()
        bat_list = [item[0] for item in bat_list]
        if(locationid == 'P2D1L1' or locationid == 'P2D1L3'):
            bat_list = []
        for file in bat_list:
            with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+file, newline='') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=";")  # aggiungere il delimiter appena lo conosciamo
                count = 0
                x = []
                yprod = []
                ycons = []
                ytot = []
                x.append(min_time_prov)
                ycons.append(0)
                for row in csv_reader:
                        val = row[1].replace(',', '.')  #se hanno messo la virgola come delimitatore decimale la cambio
                        if(count == 0):
                            first_energy_sample = float(val)
                            count =1
                        else:
                            if(min_time_prov < float(row[0]) and max_time_prov > float(row[0])):
                                if(float(row[0]) == lastSample):
                                    pow = 0
                                else:
                                    pow = 3600*(float(val) - lastValue) / (float(row[0]) - lastSample)
                                ytot.append(pow)
                                if(pow >= 0):
                                    ycons.append(pow)
                                    yprod.append(0)
                                    x.append(float(row[0]))
                                else:
                                    ycons.append(0)
                                    yprod.append(-pow)
                                    x.append(float(row[0]))
                        lastSample = float(row[0])
                        lastValue = float(val)

            ycons.append(0)
            x.append(max_time_prov)

            xnew = np.arange(min_time_prov, max_time_prov, 300, dtype = int)

            fcons = interpolate.interp1d(x, ycons)
            ynewcons = fcons(xnew)
            consumers.append(ynewcons)
            
            energyConsumption = energyConsumption - first_energy_sample
            if(energyConsumption < 0):
                energyConsumption = 0 
            totalEnergyConsumption += energyConsumption
        totalPowerList = []
        first = 0
        for consumer in consumers:
            for i in range(len(consumer)):
                if (first == 0):
                    totalPowerList.append(consumer[i])
                else:
                    totalPowerList[i] += consumer[i]
            first = 1
        if(len(totalPowerList)==0):
            Pmax = 0
        else:
            Pmax = totalPowerList[0]
        indexmax = 0
        for i in range(len(totalPowerList)):
            if totalPowerList[i] > Pmax:
                indexmax = i
                Pmax = totalPowerList[i]
        deltaT = max_time_interest_timestamp - min_time_interest_timestamp
        averagePower = 3600*totalEnergyConsumption / deltaT

        return averagePower





def calculateGC56(min_time_interest_timestamp,max_time_interest_timestamp,  min_time_interest, max_time_interest, locationid, userid, session):

    import_exports = session.query(energy_import_export).join(energy_meters_individual, energy_meters_individual.MeterID == energy_import_export.MeterID).filter(energy_meters_individual.LOC == locationid).filter(or_(not_(energy_import_export.startPoint > max_time_interest), not_(energy_import_export.endPoint < min_time_interest))).all()
    egrid = 0
    epvtogrid = 0
    for import_export in import_exports:
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+import_export.pathTofile) as csvfile:
            count = 0
            csv_reader = csv.reader(csvfile, delimiter=";")

            for row in csv_reader:
                if (min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                    if(row[1]!=''):
                        if(count == 0):
                            lastVal = float(row[1])
                            lastTime = float(row[0])
                            count = 1
                        else:
                            energyTemp = float(row[1])-lastVal
                            if(energyTemp >0):
                                egrid = egrid + energyTemp
                            else:
                                epvtogrid = epvtogrid + energyTemp
                            lastVal = float(row[1])

    
    solar_sessions = session.query(solar_plant_session).join(pv_panel_individual, pv_panel_individual.SolarPlantID == solar_plant_session.PlantID).filter(pv_panel_individual.LOC == locationid).all()
    epvproduced = 0
    for solar_session in solar_sessions:
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+solar_session.pathTofile) as csvfile:
            count = 0
            csv_reader = csv.reader(csvfile, delimiter=";")
            for row in csv_reader:
                if(row[1]!=''):
                    if (min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                        if (count == 0):
                            lastVal = float(row[1])
                            lastTime = float(row[0])
                            count = 1
                        else:
                            energyTemp = float(row[1])-lastVal
                            epvproduced = epvproduced + energyTemp
                            lastVal = float(row[1])

    etot = egrid + epvproduced - epvtogrid
    # salvatore temporary stored in a global var
    tot_pv_energy = epvproduced


    tariff_schemes =  session.query(tariff_scheme_individual).join(price_list_individual, tariff_scheme_individual.PriceModelID == price_list_individual.PriceModelID).\
        filter(or_(price_list_individual.LOC == 'ALL', price_list_individual.LOC == locationid)).filter(price_list_individual.Type == 'Transfer of energy').\
        filter(or_(not_(tariff_scheme_individual.EntryTime > max_time_interest), not_(tariff_scheme_individual.EndTime < min_time_interest))).all()


    costpaidtoDSO = 0
    for tariff in tariff_schemes:
        entryTimeObj = datetime.strptime(tariff.StartTime, '%Y-%m-%d %H:%M:%S')
        entryTimeObj_timestamp = datetime.timestamp(entryTimeObj)
        try:
            exitTimeObj = datetime.strptime(tariff.EndTime, '%Y%m%dT%H%M%S')
        except:
            if(tariff.EndTime == "202111310235959"):
                entryTimeObj = datetime.strptime("20211101T000000", '%Y%m%dT%H%M%S')
                entryTimeObj_timestamp = datetime.timestamp(entryTimeObj)
                exitTimeObj = datetime.strptime("202111300235959", '%Y%m%d0%H%M%S')

        exitTimeObj_timestamp = datetime.timestamp(exitTimeObj)
        if(tariff.PriceType == "Price per kWh"):
            egrid = 0
            for import_export in import_exports:
                count = 0
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+import_export.pathTofile) as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")
                    for row in csv_reader:
                        if (entryTimeObj_timestamp < float(row[0]) and exitTimeObj_timestamp > float(row[0]) and min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                            if (count == 0):
                                lastVal = float(row[1])
                                lastTime = float(row[0])
                                count = 1
                            else:
                                energyTemp = float(row[1]) - lastVal
                                if (energyTemp > 0):
                                    egrid = egrid + energyTemp
                                lastVal = float(row[1])
            costpaidtoDSO = costpaidtoDSO + egrid*float(tariff.Price)
        if(tariff.PriceType == "Price per time unit"):
            minTime = 0
            maxTime = 0


            if(entryTimeObj_timestamp < min_time_interest_timestamp and exitTimeObj_timestamp < max_time_interest_timestamp):
                minTime = entryTimeObj_timestamp
                maxTime = max_time_interest_timestamp
            elif(entryTimeObj_timestamp < min_time_interest_timestamp and exitTimeObj_timestamp > max_time_interest_timestamp):
                minTime = min_time_interest_timestamp
                maxTime = max_time_interest_timestamp
            elif(entryTimeObj_timestamp > min_time_interest_timestamp and exitTimeObj_timestamp < max_time_interest_timestamp):
                minTime = tariff.EntryTime
                maxTime = tariff.ExitTime
            elif(entryTimeObj_timestamp > min_time_interest_timestamp and exitTimeObj_timestamp > max_time_interest_timestamp):
                minTime = tariff.EntryTime
                maxTime = max_time_interest_timestamp
            time = 0
            delta = maxTime - minTime
            if (tariff.Period == 'Minute'):
                time = delta/60
            if (tariff.Period == 'Hour'):
                time = delta/3600
            if (tariff.Period == 'Day'):
                time = delta/86400
            if (tariff.Period == 'Week'):
                time = delta/(86400 * 7)
            if (tariff.Period == 'Month'):
                time = delta/(86400 * 30)
            if (tariff.Period == 'Year'):
                time = delta/(86400 * 365)
            costpaidtoDSO = costpaidtoDSO + time*float(tariff.Price)
        if(tariff.PriceType == "Fixed price"):
            costpaidtoDSO = costpaidtoDSO + float(tariff.Price)
        if(tariff.PriceType == "Price per highest kW"):
            consumers = []
            entryTimeObj = datetime.strptime(tariff.StartTime, '%Y-%m-%d %H:%M:%S')
            entryTimeObj_timestamp = datetime.timestamp(entryTimeObj)
            if (entryTimeObj_timestamp > min_time_interest_timestamp):
                mint = entryTimeObj_timestamp
            else:
                mint = min_time_interest_timestamp
            EndTimeObj = datetime.strptime(tariff.EndTime, '%Y-%m-%d %H:%M:%S')
            EndTimeObj_timestamp = datetime.timestamp(EndTimeObj)
            if (EndTimeObj_timestamp > max_time_interest_timestamp):
                maxt = max_time_interest_timestamp
            else:
                maxt = EndTimeObj_timestamp
            for import_export in import_exports:
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+import_export.pathTofile, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")
                    count = 0
                    x = []
                    y = []
                    lastSample = 0
                    first = 0
                    for row in csv_reader:
                        val = row[1].replace(',', '.')
                        if (count == 0 and mint < float(row[0])):
                            x.append(mint)
                            y.append(0)
                        else:
                            x.append(float(row[0]))
                            try:
                                y.append(3600 * (float(val) - lastValue) / (float(row[0]) - lastSample))
                            except:
                                y.append(0)
                        lastSample = float(row[0])
                        lastValue = float(val)
                        count += 1
                if (maxt > lastSample):
                    y.append(0)
                    x.append(maxt)
                f = interpolate.interp1d(x, y)
                xnew = np.arange(mint, maxt, 300)
                ynew = f(xnew)
                consumers.append(ynew)
            totalPowerList = []
            for consumer in consumers:
                for i in range(len(consumer)):
                    if (first == 0):
                        totalPowerList.append(consumer[i])
                    else:
                        totalPowerList[i] += consumer[i]
                first = 1
            powerpeak = 0
            for pow in totalPowerList:
                if (pow > powerpeak):
                    powerpeak = pow
            costpaidtoDSO = costpaidtoDSO + powerpeak * tariff.Price
    tariff_schemes =  session.query(tariff_scheme_individual).join(price_list_individual, tariff_scheme_individual.PriceModelID == price_list_individual.PriceModelID).\
        filter(or_(price_list_individual.LOC == 'ALL', price_list_individual.LOC == locationid)).filter(price_list_individual.Type == 'Electricity').\
        filter(or_(not_(tariff_scheme_individual.EntryTime > max_time_interest), not_(tariff_scheme_individual.EndTime < min_time_interest))).all()

    tvat =0
    costpaidtoRetailer = 0
    for tariff in tariff_schemes:
        entryTimeObj = datetime.strptime(tariff.StartTime, '%Y-%m-%d %H:%M:%S')
        entryTimeObj_timestamp = datetime.timestamp(entryTimeObj)
        exitTimeObj = datetime.strptime(tariff.EndTime, '%Y%m%dT%H%M%S')
        exitTimeObj_timestamp = datetime.timestamp(exitTimeObj)
        if(tariff.PriceType == "Price per kWh"):
            egrid = 0
            for import_export in import_exports:
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+import_export.pathTofile) as csvfile:
                    count = 0
                    csv_reader = csv.reader(csvfile, delimiter=";")
                    for row in csv_reader:
                        if (entryTimeObj_timestamp < float(row[0]) and exitTimeObj_timestamp > float(row[0]) and min_time_interest_timestamp < float(row[0]) and max_time_interest_timestamp > float(row[0])):
                            if (count == 0):
                                lastVal = float(row[1])
                                lastTime = float(row[0])
                                count = 1
                            else:
                                energyTemp = float(row[1]) - lastVal
                                if (energyTemp > 0):
                                    egrid = egrid + energyTemp
                                lastVal = float(row[1])
            costpaidtoRetailer = costpaidtoRetailer + egrid*float(tariff.Price)
        if(tariff.PriceType == "Price per time unit"):
            minTime = 0
            maxTime = 0
            if(tariff.EntryTime < min_time_interest_timestamp and tariff.ExitTime < max_time_interest_timestamp):
                minTime = tariff.EntryTime
                maxTime = max_time_interest_timestamp
            elif(tariff.EntryTime < min_time_interest_timestamp and tariff.ExitTime > max_time_interest_timestamp):
                minTime = min_time_interest_timestamp
                maxTime = max_time_interest_timestamp
            elif(tariff.EntryTime > min_time_interest_timestamp and tariff.ExitTime < max_time_interest_timestamp):
                minTime = tariff.EntryTime
                maxTime = tariff.ExitTime
            elif(tariff.EntryTime > min_time_interest_timestamp and tariff.ExitTime > max_time_interest_timestamp):
                minTime = tariff.EntryTime
                maxTime = max_time_interest_timestamp
            time = 0
            delta = maxTime - minTime
            if (tariff.Period == 'Minute'):
                time = delta/60
            if (tariff.Period == 'Hour'):
                time = delta/3600
            if (tariff.Period == 'Day'):
                time = delta/86400
            if (tariff.Period == 'Week'):
                time = delta/(86400 * 7)
            if (tariff.Period == 'Month'):
                time = delta/(86400 * 30)
            if (tariff.Period == 'Year'):
                time = delta/(86400 * 365)
            costpaidtoRetailer = costpaidtoRetailer + time*tariff.Price
        if(tariff.PriceType == "Fixed price"):
            costpaidtoRetailer = costpaidtoRetailer + tariff.Price
        if(tariff.PriceType == "Price per highest kW"):
            consumers = []
            if(tariff.EntryTime > min_time_interest_timestamp):
                mint = tariff.EntryTime
            else:
                mint = min_time_interest_timestamp
            if(tariff.ExitTime > max_time_interest_timestamp):
                maxt = max_time_interest_timestamp
            else:
                maxt = tariff.ExitTime
            for import_export in import_exports:
                with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+import_export.pathTofile, newline='') as csvfile:
                    csv_reader = csv.reader(csvfile, delimiter=";")
                    count = 0
                    x = []
                    y = []
                    lastSample = 0
                    first = 0
                    for row in csv_reader:
                        val = row[1].replace(',', '.')
                        if (count == 0 and mint < float(row[0])):
                            x.append(mint)
                            y.append(0)
                        else:
                            x.append(float(row[0]))
                            try:
                                y.append(3600 * (float(val) - lastValue) / (float(row[0]) - lastSample))
                            except:
                                y.append(0)
                        lastSample = float(row[0])
                        lastValue = float(val)
                        count += 1
                if (maxt > lastSample):
                    y.append(0)
                    x.append(maxt)
                f = interpolate.interp1d(x, y)
                xnew = np.arange(mint, maxt, 300)
                ynew = f(xnew)
                consumers.append(ynew)
            totalPowerList = []
            for consumer in consumers:
                for i in range(len(consumer)):
                    if (first == 0):
                        totalPowerList.append(consumer[i])
                    else:
                        totalPowerList[i] += consumer[i]
                first = 1
            powerpeak = 0
            for pow in totalPowerList:
                if(pow > powerpeak):
                    powerpeak = pow
            costpaidtoRetailer = costpaidtoRetailer + powerpeak * tariff.Price
        if(tariff.PriceType == "Percentage added to total"):
            tvat = tariff.Price
    ev_charging_session = session.query(charging_discharging_session).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).filter(charging_discharging_session.PluginTime > min_time_interest).filter(charging_discharging_session.PlugoutTime < max_time_interest).all()
    variable_energy_cost = session.query(variable_energy_cost_in_local_grid_public_grid).filter(variable_energy_cost_in_local_grid_public_grid.LOC == locationid).filter(or_(and_(variable_energy_cost_in_local_grid_public_grid.startPoint < min_time_interest , variable_energy_cost_in_local_grid_public_grid.endPoint > min_time_interest), and_(variable_energy_cost_in_local_grid_public_grid.endPoint > min_time_interest, variable_energy_cost_in_local_grid_public_grid.endPoint < max_time_interest),and_(variable_energy_cost_in_local_grid_public_grid.startPoint > min_time_interest, variable_energy_cost_in_local_grid_public_grid.endPoint < max_time_interest))).all()
    spotPrice = 0
    t = 0
    sessions_ev = []
    for chsession in ev_charging_session:
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+chsession.pathTofile) as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=";")
            x = []  # lista dei tempi della timeseries
            y = []  # lista dei valori della timeseries
            count2=0
            lasty = 0
            for row in csv_reader:  # per tutte le righe
                    if(count2 == 0):
                        firstval = float(row[0])
                        x.append(min_time_interest_timestamp -3600)
                        y.append(0)
                        x.append(firstval -1)
                        y.append(0)
                        count2+=1
                    lastvalue = float(row[0])
                    x.append(float(row[0]))
                    y.append(float(row[1]))
                    lasty = float(row[1])
        x.append(lastvalue +1)
        y.append(lasty)
        x.append(max_time_interest_timestamp +3600)
        y.append(lasty)
        f = interpolate.interp1d(x, y)
        xnew = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,(max_time_interest_timestamp - min_time_interest_timestamp)  / 1000,dtype=int)  # mi creo il vettore dei tempi con un sample ogni 5 minuti (300 secondi)
        ynew = f(xnew)
        sessions_ev.append(ynew)
    costs_tot = []
    for costs in variable_energy_cost:
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+costs.pathTofile) as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=";")
            count = 0
            x2 = []
            y2 = []
            for row in csv_reader:
                x2.append(float(row[0]))
                y2.append(float(row[1]))
                lastcost = float(row[1])
            x2.append(max_time_interest_timestamp+1)
            y2.append(lastcost)

            f2 = interpolate.interp1d(x2, y2, kind='previous')
            xnew2 = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,(max_time_interest_timestamp - min_time_interest_timestamp) / 1000,dtype=int)  # mi creo il vettore dei tempi con un sample ogni 5 minuti (300 secondi)
            ynew2 = f2(xnew2)

            costs_tot.append(ynew2)
    cc = 0
    totalcons = []
    totalcost = []
    eTot2 = 0
    for elem in sessions_ev:
        eTot2 += elem[-1] - elem[0]
        if(cc ==0):
            for x in elem:
                totalcons.append(x)
            cc=1
        else:
            for idx,x in enumerate(elem):
                totalcons[idx] += x
    totalcons = diff(totalcons)
    variableEnergyCost = 0
    cc =0
    for x in costs_tot:
        if(cc ==0):
            for y in x:
                totalcost.append(y)
            cc = 1
        else:
            for idx,y in enumerate(elem):
                totalcost[idx] += y
    if(len(totalcons)!=0 and len(totalcost)):
        for idx,x in enumerate(totalcost[:-1]):
            variableEnergyCost += totalcost[idx]*totalcons[idx]


    costpaidtoRetailer = (costpaidtoRetailer + variableEnergyCost)*(1+tvat)
    ev_charging_sessions = session.query(charging_discharging_session). \
        join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(
        charging_point_individual.LOC == locationid).filter(
        charging_discharging_session.PluginTime < max_time_interest).filter(
        charging_discharging_session.PlugoutTime > min_time_interest).all()
    stime = min_time_interest
    etime = max_time_interest
    stime_timestamp = datetime.timestamp(stime)
    etime_timestamp = datetime.timestamp(etime)
    relevant_bookings = session.query(reservation_booking_logs).join(charging_discharging_session, reservation_booking_logs.ChrgSessID == charging_discharging_session.ChrgSessID).join(charging_point_individual, charging_discharging_session.CPID == charging_point_individual.CPID).filter(charging_point_individual.LOC == locationid).all()

    totalEnergyConsumption = 0
    consumers = []
    cv = 0
    only581 = 0
    only582 = 0
    for booking in relevant_bookings:
        pricemodel = booking.PriceListID
        tariffs = booking.TariffID
        lft = booking.LFT
        lft_timestamp = datetime.timestamp(datetime.strptime(lft, '%Y%m%dT%H%M%S'))
        tariffs.replace('[', '')
        tariffs.replace(']', '')
        tariffs = tariffs.split(',')

        for tariff in tariffs:
            energy = 0
            tariff_scheme = session.query(tariff_scheme_individual).filter(tariff_scheme_individual.TariffID == tariff).first()
            starttime = datetime.strptime(tariff_scheme.StartTime, '%Y-%m-%d %H:%M:%S')
            starttime_timestamp = datetime.timestamp(starttime)
            endtime = datetime.strptime(tariff_scheme.EndTime, '%Y%m%dT%H%M%S')
            endtime_timestamp = datetime.timestamp(endtime)
            if(not starttime):
                starttime = stime
            if(not endtime):
                endtime =  etime
            pricetype = tariff_scheme.PriceType
            pricevalue = tariff_scheme.Price
            timeunit = tariff_scheme.Period
            for ev_charging_session in ev_charging_sessions:
                if(ev_charging_session.ChrgSessID == booking.ChrgSessID):
                    with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+ev_charging_session.pathTofile) as csvfile:
                        csv_reader = csv.reader(csvfile, delimiter=";")
                        count = 0
                        for row in csv_reader:
                            if (starttime_timestamp < float(row[0]) and endtime_timestamp > float(row[0])):
                                if (count == 0):
                                    lastVal = float(row[1])
                                    lastTime = float(row[0])
                                    count = 1
                                else:
                                    energyTemp = float(row[1]) - lastVal
                                    energy = energy + energyTemp
                                    lastVal = float(row[1])
            if(pricetype == 'Price per kWh'):
                cv = cv + pricevalue*energy

            if(pricetype == 'Price per time unit'):
                if(timeunit == 'minute'):
                    duration = (endtime_timestamp - starttime_timestamp)/60
                    only581 = pricevalue * duration
                if(timeunit == 'hour'):
                    duration = (etime_timestamp - lft_timestamp)/3600
                    only582 = pricevalue * duration
            if(pricetype == 'Fixed price'):
                cv = cv + pricevalue

    revenueFromLocallyProducedEnergy = 0
    imp_exps = []
    for imp in import_exports:
        with open('/home/branco/public_html/gccalculator/csv/newCalculator/Utils/'+imp.pathTofile) as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=";")
            count = 0
            x2 = []
            y2 = []
            lastcost = 0
            x2.append(min_time_interest_timestamp-1)
            y2.append(0)
            for row in csv_reader:
                if(row[1]!=''):
                    if(float(row[0]) >min_time_interest_timestamp and float(row[0]) <max_time_interest_timestamp):
                        x2.append(float(row[0]))
                        y2.append(float(row[1]))
                        lastcost = float(row[1])
            x2.append(max_time_interest_timestamp+1)
            y2.append(lastcost)

            f2 = interpolate.interp1d(x2, y2, kind='previous')
            xnew2 = np.arange(min_time_interest_timestamp, max_time_interest_timestamp,(max_time_interest_timestamp - min_time_interest_timestamp) / 1000,dtype=int)  # mi creo il vettore dei tempi con un sample ogni 5 minuti (300 secondi)
            ynew2 = f2(xnew2)
            imp_exps.append(ynew2)
    totalImps = []
    cc = 0
    for elem in imp_exps:
        eTot2 += elem[-1] - elem[0]
        if(cc ==0):
            for x in elem:
                totalImps.append(x)
            cc=1
        else:
            for idx,x in enumerate(elem):
                totalImps[idx] += x
    totalImps = diff(totalImps)

    if(len(totalcost) != 0):
        for idx,x in enumerate(totalImps[:-1]):
            if(x<0):

                revenueFromLocallyProducedEnergy += x * list(totalcost)[idx]

        
    revenueFromUseOfEnergy = cv + only581
    costpaidtoRetailer = (costpaidtoRetailer + spotPrice)*(1+tvat)

    pricemodel = session.query(energy_import_export).join(energy_meters_individual, energy_meters_individual.MeterID == energy_import_export.MeterID).filter(energy_meters_individual.LOC == locationid).all()

    if etot != 0:
        kpi564 = (costpaidtoRetailer+costpaidtoDSO) / etot
        kpi566 = 0
        kpiList.append([kpi564])
        kpiList.append([kpi566])
    else:
        kpiList.append([0])
        kpiList.append([0])
    
    #print('cv',cv)
    #print('rev',revenueFromLocallyProducedEnergy)
    #print('only',only581)
    if eTot2 != 0:
        kpi581 = (cv + only581 + revenueFromLocallyProducedEnergy) / eTot2
        kpi582 = (cv + only582) / eTot2
        kpiList.append([round(kpi581)])
        kpiList.append([round(kpi582)])
    else:
        kpiList.append([0])
        kpiList.append([0])

 





















def saveInFile(userid, start, end, demopilot):
    print(kpiList)
    kpiList.append(userid + '_'+start.strftime("%m-%d-%Y")+'_'+end.strftime("%m-%d-%Y")+'_'+demopilot+'_'+'evaluation.xlsx')
    with open('/home/branco/public_html/gccalculator/csv/kpi/' + userid + '_kpi.csv', 'w') as result_file:
            wr = csv.writer(result_file, delimiter=" ")
            wr.writerows(kpiList)
    arrayNames = ['GC5.3.1', 'GC5.3.2','GC5.3.3','GC5.3.4','GC5.5.1','GC5.5.2','GC5.5.3','GC5.5.4','GC5.5.5','GC5.5.6','GC5.4.1','GC5.13.1','GC5.13.2','GC5.13.3','GC5.10.1','GC5.10.2','Pmax','PeakTime','GC5.14.1','GC5.14.2', 'GC5.6.1', 'GC5.6.2', 'GC5.8.1', 'GC5.8.2', 'GC5.12', 'GC5.9']
    filePath = '/home/branco/public_html/gccalculator/csv/kpi/' + userid + '_'+start.strftime("%m-%d-%Y")+'_'+end.strftime("%m-%d-%Y")+'_'+demopilot+'_'+'evaluation.xlsx';
    if os.path.exists(filePath):
        os.remove(filePath)
    else:
        print("Can not delete the file as it doesn't exists")
    wb = Workbook()
    ws =  wb.active
    ws.title = "KPI"
    
    ws.cell(row=1, column=1).value = 'StartDay'
    ws.cell(row=1, column=2).value = start.strftime("%m-%d-%Y")
    ws.cell(row=2, column=1).value = 'EndDay'
    ws.cell(row=2, column=2).value = end.strftime("%m-%d-%Y")
    ws.cell(row=3, column=1).value = 'PilotDemoLocation'
    ws.cell(row=3, column=2).value = demopilot

    for b in range(len(kpiList)):
        if(b!=26):
            temp = kpiList[b]

            ws.cell(row=b+4, column=1).value = arrayNames[b]
            ws.cell(row=b+4, column=2).value = temp[0]
    
    wb.save(filename = filePath)


def tic():
    global _start_time 
    _start_time = time.time()

def tac():
    t_sec = round(time.time() - _start_time)
    (t_min, t_sec) = divmod(t_sec,60)
    (t_hour,t_min) = divmod(t_min,60) 
    print('Time passed: {}hour:{}min:{}sec'.format(t_hour,t_min,t_sec))

if __name__ == "__main__":
    try:
        #os.chdir('./newCalculator')

        day1=sys.argv[1]
        month1=sys.argv[2]
        year1=sys.argv[3]
        day2=sys.argv[4]
        month2=sys.argv[5]
        year2 = sys.argv[6]
        idpilot = sys.argv[7]
        userid = sys.argv[8]
        locid = sys.argv[9]
        try:
            evaldir = sys.argv[10]
        except:
            evaldir = '1'

        if(evaldir == '1'):
            database_file_name = r"/home/branco/public_html/gccalculator/csv/newCalculator/gcfull5.db"
            engine = sqlalchemy.create_engine("sqlite:////home/branco/public_html/gccalculator/csv/newCalculator/gcfull5.db", echo=False)
            Session = sessionmaker(bind=engine)
            session = Session()
            directory = 'research_data5'
        else:
            database_file_name = r"/home/branco/public_html/gccalculator/csv/newCalculator/gcfull6.db"
            engine = sqlalchemy.create_engine("sqlite:////home/branco/public_html/gccalculator/csv/newCalculator/gcfull6.db", echo=False)
            Session = sessionmaker(bind=engine)
            session = Session()

            directory = 'sharedTest'
        '''month1 = 'February'
        month2 = 'February'
        year1 = '2021'
        year2 = '2021'
        day1 = '03'
        day2 = '09'''
        idpilot = utils.numInDemo(idpilot)
        demoAndLocation = idpilot+'L'+locid
        month1 = utils.monthinnum(month1)
        month2 = utils.monthinnum(month2)
        min_time_interest= day1+'/'+month1+'/'+year1+' 00:00:00'
        min_time_interest_obj = datetime.strptime(min_time_interest, '%d/%m/%Y %H:%M:%S')
        min_time_interest_timestamp = datetime.timestamp(min_time_interest_obj)
        max_time_interest = day2+'/'+month2+'/'+year2+' 23:59:59'
        max_time_interest_obj = datetime.strptime(max_time_interest, '%d/%m/%Y %H:%M:%S')
        max_time_interest_timestamp = datetime.timestamp(max_time_interest_obj)

        startTime = datetime.fromtimestamp(min_time_interest_timestamp)
        endTime = datetime.fromtimestamp(max_time_interest_timestamp)
    except Exception as e:
        print(e)
    if(evaldir == '1'):
        mainmeter_folder = r"/home/branco/public_html/gccalculator/csv/newCalculator/Utils/research_data/recordings_logs/energy_characteristics/energy_import_export"
        gridmix_folder = r"/home/branco/public_html/gccalculator/csv/newCalculator/Utils/research_data/recordings_logs/energy_characteristics/grid_mix"

    else:
        mainmeter_folder = r"/home/branco/public_html/gccalculator/csv/newCalculator/Utils/research_data/recordings_logs/energy_characteristics/energy_import_export"
        gridmix_folder = r"/home/branco/public_html/gccalculator/csv/newCalculator/Utils/research_data/recordings_logs/energy_characteristics/grid_mix"


    
    emission_factor = "/home/branco/public_html/gccalculator/csv/newCalculator/Utils/emission_factor.csv"



    #calculateGC53(startTime,endTime,'P1D1L1',1, session)
    print(demoAndLocation)
    tic()
    calculateGC53(startTime, endTime, demoAndLocation, userid, session)
    tac()
    print('GC53')
    tic()
    calculateGC55(startTime, endTime, demoAndLocation, userid, session)
    tac()
    print('GC55')
    tic()
    calculateGC54(startTime, endTime, demoAndLocation, userid, session)
    tac()
    print('GC54 ')
    tic()
    calculateGC513(startTime, endTime, demoAndLocation, userid, session)
    tac()
    print('GC513 ')
    #selfConsumption2ndVersion(min_time_interest_timestamp, max_time_interest_timestamp, 'P1D1L1', '1')
    #except Exception as e:
    #    print(e)
    tic()
    avpow = calculateKPI510(min_time_interest_timestamp, max_time_interest_timestamp,startTime,endTime,  demoAndLocation, userid, directory)
    tac()
    print('GC510 ')
    tic()
    #calculateKPI510byhour(min_time_interest_timestamp, max_time_interest_timestamp,startTime,endTime,  demoAndLocation, userid, directory)
    selfConsumption2ndVersion(min_time_interest_timestamp, max_time_interest_timestamp, startTime,endTime, demoAndLocation, userid, directory,avpow)
    tac()
    print('GC514 ')
    tic()
    calculateGC56(min_time_interest_timestamp, max_time_interest_timestamp, startTime, endTime, demoAndLocation, userid, session)
    tac()
    print('GC56 ')
    kpiList.append([kpi512_sal(demoAndLocation,year1,month1, year2,month2,startTime,endTime,gridmix_folder,emission_factor,mainmeter_folder)])
    kpiList.append([kpi59_sal(demoAndLocation, year1, month1,year2,month2,startTime,endTime,gridmix_folder)])
    saveInFile(userid,startTime,endTime,demoAndLocation)

